# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
from pathlib import Path
import subprocess



def MovementReportAccount(objJson : dict):
    
    celdaNombre             = "B3"
    celdaCorreo             = "B4"
    celdaCuenta             = "B5"
    celdaFechaCreacion      = "J3"
    celdaFechaRango         = "J4"
    celdaMovimiento         = "A"
    incrementoMov           = 8
    incrementoNuevasFilas   = 0

    # Creo copia de reporte
    DIR_MOVEMENT_REPORT             = "TEMPLATES/mobile/MovementReport/Reporte_de_movimiento_Cuenta.xlsx"
    DIR_TMP_MOVEMENT_REPORT         = "TMP/mobile/MovementReport/Reporte_de_movimiento_Cuenta_" + str(objJson["correo"]) + "_" + str(
        datetime.today().strftime("%Y-%m-%d_%H-%M-%S")) + ".xlsx"
    PATH_DIR_MOVEMENT_REPORT        = Path("../../" + str(DIR_MOVEMENT_REPORT)).absolute()
    PATH_DIR_MOVEMENT_REPORT        = str(PATH_DIR_MOVEMENT_REPORT).replace("../../", "")
    PATH_DIR_TMP_MOVEMENT_REPORT    = Path("../../" + str(DIR_TMP_MOVEMENT_REPORT)).absolute()
    PATH_DIR_TMP_MOVEMENT_REPORT    = str(PATH_DIR_TMP_MOVEMENT_REPORT).replace("../../", "")
    cmdExt                          = "cp " + str(PATH_DIR_MOVEMENT_REPORT) + " " + str(PATH_DIR_TMP_MOVEMENT_REPORT)
    output                          = subprocess.check_output(cmdExt, shell=True, universal_newlines=True)
    output                          = output.strip()
    """
    reporteTmp      = "Reporte_de_movimiento_Cuenta_"+str(objJson["correo"])+"_"+str(datetime.today().strftime("%Y-%m-%d_%H-%M-%S"))+".xlsx"
    #fpath			= Path("./MANAGEMENT/EncryptDecrypt/openssl.php").absolute()
    #fpath			= str(fpath).replace("./","")
    cmdExt			= "cp ./Reporte_de_movimiento_Cuenta.xlsx ./"+str(reporteTmp)
    output			= subprocess.check_output(cmdExt, shell=True, universal_newlines=True)
    output			= output.strip()
    """
    
    #wb = Workbook()
    wb  = load_workbook( str(PATH_DIR_TMP_MOVEMENT_REPORT) )
    ws = wb.active

    # Asigno nombre
    ws[celdaNombre]         = objJson["nombre"]

    # Asigno correo
    ws[celdaCorreo]         = objJson["correo"]

    # Asungo cuenta
    ws[celdaCuenta]         = "No. de Cuenta: " + str(objJson["tipo"]["cuenta"]["valor"])

    # Asigno fecha creación
    ws[celdaFechaCreacion]  = datetime.now()

    # Asigno rango de fecha
    ws[celdaFechaRango]     = "Periodo de reporte: Del "+str(objJson["fechaini"])+" al "+str(objJson["fechafin"])

    # Agrego filas necesarias
    """
    filasNuevas     = 0
    tamDeRegistros  = len( objJson["tipo"]["cuenta"]["movimientos"])
    if tamDeRegistros >= 49:
        filasNuevas = len( objJson["tipo"]["cuenta"]["movimientos"]) - 48
    """

    # Asigno movimientos
    formato     = Font(name="Arial", size=12, color="00000000")
    alineacion  = Alignment(horizontal="center", vertical="center")
    ultimaFilaDelMovimiento = 0
    if len(objJson["tipo"]["cuenta"]["movimientos"]) >= 1:

        for movimiento in objJson["tipo"]["cuenta"]["movimientos"]:

            #moduloFilasNuevas   =  incrementoNuevasFilas % 48
            #if tamDeRegistros >= 49 and moduloFilasNuevas == 0:
            #    ws.insert_rows(incrementoMov, filasNuevas)

            ws[ "A" + str(incrementoMov) ]              = movimiento["fecha"]
            ws[ "A" + str(incrementoMov) ].font         = formato
            ws[ "A" + str(incrementoMov) ].alignment    = alineacion

            ws.merge_cells( "B" + str(incrementoMov) + ":E" + str(incrementoMov) )
            ws[ "B" + str(incrementoMov) ]              = movimiento["nombre_beneficiario"]
            ws[ "B" + str(incrementoMov) ].font         = formato
            ws[ "B" + str(incrementoMov) ].alignment    = alineacion

            ws.merge_cells( "F" + str(incrementoMov) + ":G" + str(incrementoMov) )
            ws[ "F" + str(incrementoMov) ]              = movimiento["cta_beneficiario"]
            ws[ "F" + str(incrementoMov) ].font         = formato
            ws[ "F" + str(incrementoMov) ].alignment    = alineacion

            ws[ "H" + str(incrementoMov) ]              = movimiento["monto"]
            ws[ "H" + str(incrementoMov) ].font         = formato
            ws[ "H" + str(incrementoMov) ].alignment    = alineacion

            ws.merge_cells( "I" + str(incrementoMov) + ":J" + str(incrementoMov) )
            ws[ "I" + str(incrementoMov) ]              = movimiento["concepto"]
            ws[ "I" + str(incrementoMov) ].font         = formato
            ws[ "I" + str(incrementoMov) ].alignment    = alineacion

            ws.merge_cells( "K" + str(incrementoMov) + ":L" + str(incrementoMov) )
            ws[ "K" + str(incrementoMov) ]              = movimiento["referencia"]
            ws[ "K" + str(incrementoMov) ].font         = formato
            ws[ "K" + str(incrementoMov) ].alignment    = alineacion

            incrementoMov   += 1
            incrementoNuevasFilas   += 1
            ultimaFilaDelMovimiento = incrementoMov
    else:
            ultimaFilaDelMovimiento = 8

    # PIE DE PAGINA: Ajustando incremento de filas para separarlo del ultimo movimiento
    deltaFilas  = ultimaFilaDelMovimiento + 2
    # PIE DE PAGINA: Combino celdas para pie de pagina (fila1)
    ws.row_dimensions[ ultimaFilaDelMovimiento ].height    = 22.20
    ws.merge_cells( "A" + str(ultimaFilaDelMovimiento) + ":L" + str(ultimaFilaDelMovimiento) )
    # PIE DE PAGINA: Combino celdas para pie de pagina (fila2)
    ultimaFilaDelMovimiento += 1
    ws.row_dimensions[ ultimaFilaDelMovimiento ].height = 64.20

    pieDePagina_fila2_cellA_formato     = Font(name="Arial", size=12, color="00FFFFFF")
    pieDePagina_fila2_cellA_alineacion  = Alignment(horizontal="left", vertical="center")
    pieDePagina_fila2_colorDeFondo      = PatternFill(start_color="00022E92", end_color="00022E92", fill_type = "solid")
    ws[ "A" + str(ultimaFilaDelMovimiento) ].value     = ""
    ws[ "A" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellA_formato
    ws[ "A" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellA_alineacion
    ws[ "A" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    pieDePagina_fila2_cellB_formato     = Font(name="Arial", size=12, color="00FFFFFF")
    pieDePagina_fila2_cellB_alineacion  = Alignment(horizontal="center", vertical="center")
    ws.merge_cells( "B" + str(ultimaFilaDelMovimiento) + ":E" + str(ultimaFilaDelMovimiento) )
    ws[ "B" + str(ultimaFilaDelMovimiento) ].value     = "T. (+52) 554 170 4129\nT. (+52) 556 8278522"
    ws[ "B" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellB_formato
    ws[ "B" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellB_alineacion
    ws[ "B" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    ws[ "F" + str(ultimaFilaDelMovimiento) ].value     = ""
    ws[ "F" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellA_formato
    ws[ "F" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellA_alineacion
    ws[ "F" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    pieDePagina_fila2_cellG_formato     = Font(name="Calibri", size=12, color="00FFFFFF")
    ws.merge_cells( "G" + str(ultimaFilaDelMovimiento) + ":H" + str(ultimaFilaDelMovimiento) )
    ws[ "G" + str(ultimaFilaDelMovimiento) ].value      = "E. info@polipay.mx"
    ws[ "G" + str(ultimaFilaDelMovimiento) ].font       = pieDePagina_fila2_cellG_formato
    ws[ "G" + str(ultimaFilaDelMovimiento) ].alignment  = pieDePagina_fila2_cellB_alineacion
    ws[ "G" + str(ultimaFilaDelMovimiento) ].fill       = pieDePagina_fila2_colorDeFondo

    ws[ "I" + str(ultimaFilaDelMovimiento) ].value     = ""
    ws[ "I" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellA_formato
    ws[ "I" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellA_alineacion
    ws[ "I" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    ws.column_dimensions["J"].width  = 58.73
    pieDePagina_fila2_cellJ_formato     = Font(name="Arial", size=11, color="00FFFFFF")
    ws[ "J" + str(ultimaFilaDelMovimiento) ].value     = " D. CP. 45129, Zapopan, Jalisco.\nD. CP. 07750, Ciudad de México, CDMX."
    ws[ "J" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellJ_formato
    ws[ "J" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellB_alineacion
    ws[ "J" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    ws[ "K" + str(ultimaFilaDelMovimiento) ].value     = ""
    ws[ "K" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellA_formato
    ws[ "K" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellA_alineacion
    ws[ "K" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    ws[ "L" + str(ultimaFilaDelMovimiento) ].value     = ""
    ws[ "L" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila2_cellA_formato
    ws[ "L" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila2_cellA_alineacion
    ws[ "L" + str(ultimaFilaDelMovimiento) ].fill      = pieDePagina_fila2_colorDeFondo

    # PIE DE PAGINA: Combino celdas para pie de pagina (fila3) y agrego leyenda
    pieDePagina_fila3_leyenda   = "Los recursos de los Usuarios en las operaciones realizadas con Polipay  no se encuentran garantizados por ninguna autoridad. Los fondos de pago electrónico no generan rendimientos o beneficios monetarios por los saldos acumulados en los mismos. Polipay  recibe consultas, reclamaciones o aclaraciones, en su Unidad Especializada de Atención a Usuarios, por correo electrónico a contacto@polipay.com . En el caso de no obtener una respuesta satisfactoria, podrá acudir a la Comisión Nacional para la Protección y Defensa de los Usuarios de Servicios Financieros a través de su página web: https//gob.mx/condusef o al número telefónico 5553400999."

    ultimaFilaDelMovimiento += 1
    ws.row_dimensions[ ultimaFilaDelMovimiento ].height    = 39.00

    pieDePagina_fila3_formato       = Font(name="Arial", size=10, color="00000000")
    pieDePagina_fila3_alineacion    = Alignment(horizontal="left", vertical="center")
    ws.merge_cells( "A" + str(ultimaFilaDelMovimiento) + ":L" + str(ultimaFilaDelMovimiento) )
    ws[ "A" + str(ultimaFilaDelMovimiento) ].value     = pieDePagina_fila3_leyenda
    ws[ "A" + str(ultimaFilaDelMovimiento) ].font      = pieDePagina_fila3_formato
    ws[ "A" + str(ultimaFilaDelMovimiento) ].alignment = pieDePagina_fila3_alineacion

    # Save the file
    wb.save( str(PATH_DIR_TMP_MOVEMENT_REPORT) )

    return str(PATH_DIR_TMP_MOVEMENT_REPORT)

