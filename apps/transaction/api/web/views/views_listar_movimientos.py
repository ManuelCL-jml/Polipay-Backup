import abc
import os
import datetime as dt
import mimetypes
from abc import ABC

import openpyxl

from typing import Dict, List, Any, ClassVar, Union

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q
from django.db.models.query_utils import PathInfo

from django.db.transaction import atomic
from django.http.response import HttpResponse, FileResponse
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
import datetime
from MANAGEMENT.Standard.errors_responses import MyHttpError
from MANAGEMENT.Utils.utils import remove_asterisk
from apps.logspolipay.manager import RegisterLog
from apps.transaction.api.web.views.Transacciones_views import Transacciones
from apps.transaction.exc import ParamsNotProvided, ParamsRaiseException
from apps.users.messages import pruebas

from rest_framework.generics import ListAPIView, RetrieveAPIView, UpdateAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework import status
from rest_framework.viewsets import GenericViewSet

from dateutil.relativedelta import relativedelta

from operator import itemgetter

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from polipaynewConfig.exceptions import ErrorsList
from polipaynewConfig.exceptions import *
from apps.transaction.management import *
from apps.users.management import get_data_empresa, get_id_cuenta_eje
from apps.transaction.models import *
from apps.users.models import cuenta, persona, domicilio


# (ChrGil 2021-10-13) Depenediendo del valor del parametro, se realizara una tarea distintan
def to_dict_query_params(query_params) -> Dict:
    data: Dict[str, Any] = {}
    for key, value in query_params.items():
        if value == 'null':
            if key == 'date_1':
                new_date = dt.datetime.now() - dt.timedelta(days=91)
                data[key] = new_date
                continue

            if key == 'date_2':
                data[key] = dt.datetime.now()
                continue

            data[key] = ''
        if value != 'null':
            data[key] = value

    data.pop('size')
    return data


class GetInfoCompany:
    info_account: ClassVar[Dict[str, Any]]

    def __init__(self, admin: persona, **kwargs):
        self._admin = admin
        self._razon_social_id = kwargs.get('razon_social_id', None)
        self._get_info_account_cuenta_eje()

        if not self._razon_social_id:
            raise ParamsNotProvided('Operación prohibida, debes de enviar el id del una persona moral')

        if self.info_account.get('rel_cuenta_prod_id') == 3:
            if not self._exists_cost_center:
                raise ValueError('Centro de costos no valido o no existe')

            self._get_info_account_cost_center()

    @property
    def _get_cuenta_eje(self) -> int:
        return get_id_cuenta_eje(self._admin.get_only_id())

    def _get_info_account_cuenta_eje(self):
        self.info_account = cuenta.objects.filter(
            persona_cuenta_id=self._get_cuenta_eje,
            is_active=True
        ).values('id', 'rel_cuenta_prod_id', 'persona_cuenta_id', 'cuenta', 'cuentaclave').first()

    @property
    def _exists_cost_center(self) -> bool:
        return grupoPersona.objects.select_related('empresa', 'person', 'relacion_grupo').filter(
            empresa_id=self.info_account.get('persona_cuenta_id'),
            person_id=self._razon_social_id,
            relacion_grupo_id=5
        ).exists()

    def _get_info_account_cost_center(self):
        self.info_account = cuenta.objects.select_related('persona_cuenta').filter(
            persona_cuenta_id=self._razon_social_id,
            is_active=True
        ).values('id', 'cuenta', 'cuentaclave').first()


class ListData(ABC):
    defaul_size: ClassVar[int] = 5
    data: ClassVar[Union[List[Dict[str, Any]], Dict[str, Any]]]


# (ChrGil 2022-02-13) Listar Egresos dependiendo del producto
class ClassListEgresos(ListData):
    list_egresos: ClassVar[List[Dict[str, Any]]]

    def __init__(self, razon_social: GetInfoCompany, **kwargs):
        self.size = kwargs.get('size', self.defaul_size)
        self._start_date = kwargs.get('date_1', dt.date.today() - dt.timedelta(days=91))
        self._end_date = kwargs.get('date_2', dt.date.today())
        self._emisor = razon_social
        self.list_egresos = [self.agrega_tipo_momiviento(**egreso) for egreso in self._list_egreso]

    @property
    def _list_egreso(self) -> List[Dict[str, Any]]:
        return transferencia.objects.filter(
            Q(fecha_creacion__date__gte=self._start_date) &
            Q(fecha_creacion__date__lte=self._end_date)
        ).filter(
            Q(cuenta_emisor__icontains=self._emisor.info_account.get('cuenta')) |
            Q(cuenta_emisor__icontains=self._emisor.info_account.get('cuentaclave'))
        ).filter(
            status_trans_id__in=[1, 9, 7]
        ).values(
            'id',
            'nombre_beneficiario',
            'cta_beneficiario',
            'clave_rastreo',
            'concepto_pago',
            'monto',
            'saldo_remanente',
            'fecha_creacion',
            'status_trans_id'
        ).exclude(
            tipo_pago_id=10
        ).order_by('-fecha_creacion')

    @staticmethod
    def agrega_tipo_momiviento(**kwargs):
        return {
            'id': kwargs.get('id'),
            'nombre_beneficiario': remove_asterisk(kwargs.get('nombre_beneficiario')),
            'cta_beneficiario': kwargs.get('cta_beneficiario'),
            'clave_rastreo': kwargs.get('clave_rastreo'),
            'concepto_pago': kwargs.get('concepto_pago'),
            'monto': kwargs.get('monto'),
            'saldo_remanente': kwargs.get('saldo_remanente'),
            'fecha_creacion': kwargs.get('fecha_creacion'),
            'tipo_movimiento': kwargs.get('tipo_movimiento', 'Egreso'),
            'status_trans_id': kwargs.get('status_trans_id')
        }


class ListEgresosRazonSocial(ListAPIView):
    pagination_class = PageNumberPagination

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        try:
            admin: persona = request.user
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            emisor = GetInfoCompany(admin, **data)
            egreso = ClassListEgresos(emisor, **data)
            self.pagination_class.page_size = egreso.size
            return self.get_paginated_response(self.paginate_queryset(egreso.list_egresos))
        except (ObjectDoesNotExist, ValueError, ParamsRaiseException) as e:
            err = MyHttpError('Su cuenta no esta asociada a una persona moral', real_error=str(e))
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)


# (ChrGil 2022-02-13) Listar Ingresos dependiendo del producto
class ClassListIngreso(ListData):
    list_ingreso: ClassVar[List[Dict[str, Any]]]

    def __init__(self, razon_social: GetInfoCompany, **kwargs):
        self.size = kwargs.get('size', self.defaul_size)
        self._start_date = kwargs.get('date_1', dt.date.today() - dt.timedelta(days=91))
        self._end_date = kwargs.get('date_2', dt.date.today())
        self._razon_social_info = razon_social.info_account
        self.list_ingreso = [self.agrega_tipo_momiviento(**ingreso) for ingreso in self._list_ingreso]

    @property
    def _list_ingreso(self) -> List[Dict[str, Any]]:
        return transferencia.objects.filter(
            Q(fecha_creacion__date__gte=self._start_date) &
            Q(fecha_creacion__date__lte=self._end_date)
        ).filter(
            Q(cta_beneficiario__icontains=self._razon_social_info.get('cuenta')) |
            Q(cta_beneficiario__icontains=self._razon_social_info.get('cuentaclave'))
        ).values(
            'id',
            'nombre_beneficiario',
            'cta_beneficiario',
            'clave_rastreo',
            'concepto_pago',
            'monto',
            'saldo_remanente_beneficiario',
            'fecha_creacion',
            'status_trans_id',
        ).exclude(
            tipo_pago_id=10
        ).order_by('-fecha_creacion')

    @staticmethod
    def agrega_tipo_momiviento(**kwargs):
        return {
            'id': kwargs.get('id'),
            'nombre_beneficiario': remove_asterisk(kwargs.get('nombre_beneficiario')),
            'cta_beneficiario': kwargs.get('cta_beneficiario'),
            'clave_rastreo': kwargs.get('clave_rastreo'),
            'concepto_pago': kwargs.get('concepto_pago'),
            'monto': kwargs.get('monto'),
            'saldo_remanente': kwargs.get('saldo_remanente_beneficiario'),
            'fecha_creacion': kwargs.get('fecha_creacion'),
            'tipo_movimiento': kwargs.get('tipo_movimiento', 'Ingreso'),
            'status_trans_id': kwargs.get('status_trans_id')
        }


class ListIngresosRazonSocial(ListAPIView):
    pagination_class = PageNumberPagination

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        try:
            admin: persona = request.user
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            emisor = GetInfoCompany(admin, **data)
            egreso = ClassListIngreso(emisor, **data)
            self.pagination_class.page_size = egreso.size
            return self.get_paginated_response(self.paginate_queryset(egreso.list_ingreso))
        except (ObjectDoesNotExist, ValueError, ParamsRaiseException) as e:
            err = MyHttpError('Su cuenta no esta asociada a una persona moral', real_error=str(e))
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)


# (ChrGil 2022-02-13) Listar ingresos y egresos dependiendo del producto
class ClassListIngresoEgreso(ListData):
    list_ingreso_egreso: ClassVar[List[Dict[str, Any]]]
    _cuenta: ClassVar[str]
    _cuentaclave: ClassVar[str]

    def __init__(self, razon_social: GetInfoCompany, **kwargs):
        self.size = kwargs.get('size', self.defaul_size)
        self._start_date = kwargs.get('date_1', dt.date.today() - dt.timedelta(days=91))
        self._end_date = kwargs.get('date_2', dt.date.today())
        self._razon_social_info = razon_social.info_account
        self._cuenta = razon_social.info_account.get('cuenta')
        self._cuentaclave = razon_social.info_account.get('cuentaclave')
        self.list_ingreso_egreso = self.agrega_tipo_movimiento

    @property
    def _list_ingreso(self) -> List[Dict[str, Any]]:
        return transferencia.objects.filter(
            Q(fecha_creacion__date__gte=self._start_date) &
            Q(fecha_creacion__date__lte=self._end_date)
        ).filter(
            Q(cta_beneficiario__icontains=self._razon_social_info.get('cuenta')) |
            Q(cta_beneficiario__icontains=self._razon_social_info.get('cuentaclave')) |
            Q(cuenta_emisor__icontains=self._razon_social_info.get('cuenta')) |
            Q(cuenta_emisor__icontains=self._razon_social_info.get('cuentaclave'))
        ).filter(
            status_trans_id__in=[1, 9, 7]
        ).values(
            'id',
            'nombre_beneficiario',
            'cta_beneficiario',
            'cuenta_emisor',
            'clave_rastreo',
            'concepto_pago',
            'monto',
            'saldo_remanente',
            'saldo_remanente_beneficiario',
            'fecha_creacion',
            'date_modify',
            'status_trans_id'
        ).exclude(
            tipo_pago_id=10
        ).order_by('-fecha_creacion')

    @property
    def agrega_tipo_movimiento(self) -> List[Dict[str, Any]]:
        list_ingreso_egreso = list(self._list_ingreso).copy()

        for i in range(0, len(list_ingreso_egreso)):
            if list_ingreso_egreso[i].get("cta_beneficiario") == "5858587556":
                continue

            if list_ingreso_egreso[i].get('cta_beneficiario') == self._cuenta:
                list_ingreso_egreso[i]['tipo_movimiento'] = 'Ingreso'
                # list_ingreso_egreso[i]['saldo_remanente'] = None

            if list_ingreso_egreso[i].get('cta_beneficiario') == self._cuentaclave:
                list_ingreso_egreso[i]['tipo_movimiento'] = 'Ingreso'
                # list_ingreso_egreso[i]['saldo_remanente'] = None

            if list_ingreso_egreso[i].get('cuenta_emisor') == self._cuenta:
                list_ingreso_egreso[i]['tipo_movimiento'] = 'Egreso'
                list_ingreso_egreso[i]['saldo_remanente_beneficiario'] = None

            if list_ingreso_egreso[i].get('cuenta_emisor') == self._cuentaclave:
                list_ingreso_egreso[i]['tipo_movimiento'] = 'Egreso'
                list_ingreso_egreso[i]['saldo_remanente_beneficiario'] = None

        return list_ingreso_egreso


class ListIngresosEgresosRazonSocial(ListAPIView):

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        try:
            admin: persona = request.user
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            emisor = GetInfoCompany(admin, **data)
            egreso = ClassListIngresoEgreso(emisor, **data)
            self.pagination_class.page_size = egreso.size
            return self.get_paginated_response(self.paginate_queryset(egreso.list_ingreso_egreso))
        except (ObjectDoesNotExist, ValueError, ParamsRaiseException) as e:
            err = MyHttpError('Su cuenta no esta asociada a una persona moral', real_error=str(e))
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)


# (ChrGil 2021-10-19) Se implementa la función to_dict_query_params
class GeneralList(ListAPIView):
    pagination_class = PageNumberPagination

    def list_movimientos(self, queryset: List, cuenta_emisor: str = None) -> None:
        return None

    def filter_data(self, cuenta_beneficiario: str, date_1, date_2) -> None:
        return None

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        self.pagination_class.page_size = request.query_params['size']
        data = to_dict_query_params(request.query_params)
        log.json_request(data)

        data_company: Dict = get_data_empresa(request.user.get_only_id())
        get_cuenta: str = cuenta.objects.get(persona_cuenta_id=data_company['id']).get_cuenta()

        queryset = self.filter_data(get_cuenta, **data)
        page = self.paginate_queryset(self.list_movimientos(queryset, get_cuenta))
        return self.get_paginated_response(page)


# (ChrGil 2021-10-19)
class ListIngresos(GeneralList):
    def list_movimientos(self, queryset: List, cuenta_emisor: str = None) -> List:
        for query in queryset:
            query['tipo_movimiento'] = 'Ingreso'
        return queryset

    def filter_data(self, cuenta_beneficiario: str, date_1, date_2) -> List:
        return transferencia.filter_transaction.list_income(cuenta_beneficiario, date_1, date_2)


class ListEgresos(GeneralList):
    def list_movimientos(self, queryset: List, cuenta_emisor: str = None) -> List:
        for query in queryset:
            query['tipo_movimiento'] = 'Egreso'
        return queryset

    def filter_data(self, cuenta_emisor: str, date_1, date_2) -> List:
        return transferencia.filter_transaction.list_expenses(cuenta_emisor, date_1, date_2)


# (ChrGil 2021-10-19) Listado de ingreso y egresos ordenado por fecha
class ListIngresoEngresoV2(GeneralList):
    def list_movimientos(self, queryset: List, cuenta_emisor: str = None) -> List:
        for query in queryset:
            if query['cta_beneficiario'] == cuenta_emisor:
                query['tipo_movimiento'] = 'Ingreso'
            else:
                query['tipo_movimiento'] = 'Egreso'

        return queryset

    def filter_data(self, cuenta_emisor: str, date_1, date_2) -> List:
        return transferencia.filter_transaction.list_income_and_expenses(cuenta_emisor, date_1, date_2)


class pdfmasivas(GenericViewSet):
    serializer_class = None

    def delete(self, request):
        try:
            username = request.query_params["user"]
            os.remove("TMP/web/movimientos-DispMassiva" + username + ".pdf")
            return Response({"status": "pdf eliminado"}, status=status.HTTP_200_OK)
        except:
            return Response({"status": "no se encontro el pdf"}, status=status.HTTP_400_BAD_REQUEST)

    @method_decorator(cache_page(60 * 0.1))
    def list(self, requets):
        try:
            username = self.request.query_params["user"]
            filename = 'TMP/web/movimientos-DispMassiva' + username + ".pdf"
            filepath = filename
            path = open(filepath, 'r')
            mime_type, _ = mimetypes.guess_type(filepath)
            response = FileResponse(open(filename, 'rb'))
            response['Content-Disposition'] = "attachment; filename=%s" % filename
            return response
        except:
            return Response({"status": "no se encontro el pdf"}, status=status.HTTP_400_BAD_REQUEST)


class CrearPDFMasivas(ListAPIView):
    permission_classes = ()

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        errores = []
        masivoId = request.query_params["masId"]
        queryset = transferencia.objects.filter(masivo_trans_id=masivoId)
        if queryset:
            pass
        else:
            errores.append({"field": "masId", "data": masivoId, "message": "transferencia masiva no encontrada"})
        username = request.query_params["user"]
        if persona.objects.filter(username=username):
            PDFMasivas(queryset, username, masivoId)
            return Response({"status": "pdf creado"}, status=status.HTTP_200_OK)
        else:
            errores.append({"field": "user", "data": username, "message": "username no encontrado"})
        return Response({"Error": {"code": ["400"]}, "status": ["ERROR"], "detail": [errores]})


# (2021-12-20 Jose) End point pata listar/filtro de movimientos (egresos)
class MovimientosEgresos(GenericViewSet):
    permission_classes = ()
    pagination_class = PageNumberPagination

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request):
        error = []
        meses = request.query_params["Meses"]
        fecha_inicio = request.query_params["FechaInicio"]
        fecha_final = request.query_params["FechaFinal"]

        size = NumInt(size=self.request.query_params["size"])
        self.pagination_class.page_size = size

        if meses == "Null" and fecha_inicio == "Null" and fecha_final == "Null":  # ----------> Query sin filtro
            egresos = transferencia.objects.filter(cuenta_emisor=request.query_params["NumeroCuenta"]).order_by(
                "-fecha_creacion")

        else:
            if meses != "Null" and fecha_inicio != "Null" and fecha_final != "Null":
                error.append({"field": "Meses,FechaInicio,FechaFinal", "data": "",
                              "message": "No se puede buscar por meses y fechas al mismo tiempo"})

            if meses != "Null":
                if int(meses) > 2:
                    error.append({"field": "Meses", "data": meses,
                                  "message": "El rango de busqueda por mes no puede ser mayor que 2"})
                else:  # -----------> Query con filtro de meses (1 mes o 2 meses)
                    Fecha_actual = datetime.date.today()
                    Fecha_hasta_hora = str(Fecha_actual) + " 23:59:59"
                    fecha_final = datetime.datetime.strptime(Fecha_hasta_hora, "%Y-%m-%d %H:%M:%S")
                    if int(meses) == 1:
                        fecha_inicio = Fecha_actual.replace(day=1)
                    else:
                        fecha_inicio = Fecha_actual - relativedelta(months=int(1))
                    Fecha_inicio_hora = str(fecha_inicio) + " 00:00:00"
                    fecha_inicio = datetime.datetime.strptime(Fecha_inicio_hora, "%Y-%m-%d %H:%M:%S")

            if meses == "Null" and fecha_inicio != "Null" or "" and fecha_final != "Null" or "":  # ----------> Query con filtro de fechas
                FechaDesdeHora = str(fecha_inicio) + " 00:00:00"
                FechaHastaHora = str(fecha_final) + " 23:59:59"
                fecha_inicio = datetime.datetime.strptime(FechaDesdeHora, "%Y-%m-%d %H:%M:%S")
                fecha_final = datetime.datetime.strptime(FechaHastaHora, "%Y-%m-%d %H:%M:%S")

            egresos = transferencia.objects.filter(cuenta_emisor=request.query_params["NumeroCuenta"],
                                                   fecha_creacion__gte=fecha_inicio,
                                                   fecha_creacion__lte=fecha_final).order_by("-fecha_creacion")

        if error:
            MensajeError(error)
        n = 0
        list = []
        for datos in egresos:
            tipo = tipo_transferencia.objects.get(id=datos.tipo_pago_id)
            fecha, hora = str(datos.fecha_creacion).split(" ")
            origen = "Individual"
            if datos.masivo_trans_id != None:
                origen = "Masiva"
            diccionario = {
                "id": datos.id,
                "Beneficiario": datos.nombre_beneficiario,
                "Monto": f'-${datos.monto}',
                "FechaOperacion": fecha.replace("-", "/ "),
                "Tipo": tipo.nombre_tipo,
                "Origen": origen
            }
            list.append(diccionario)
            n = n + 1
            if n == size:
                break
        page = self.paginate_queryset(list)
        return Response(page, status=status.HTTP_200_OK)


# (2021-12-20 Jose) End point pata listar/filtro de movimientos (ingresos)
class MovimientosIngresos(GenericViewSet):
    pagination_class = PageNumberPagination
    permission_classes = ()

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request):
        error = []
        meses = request.query_params["Meses"]
        fecha_inicio = request.query_params["FechaInicio"]
        fecha_final = request.query_params["FechaFinal"]

        size = NumInt(size=self.request.query_params["size"])
        self.pagination_class.page_size = size

        if meses == "Null" and fecha_inicio == "Null" and fecha_final == "Null":  # ----------> Query sin filtro

            ingresos = transferencia.objects.filter(cta_beneficiario=request.query_params["NumeroCuenta"]).order_by(
                "-fecha_creacion")

        else:

            if meses != "Null" and fecha_inicio != "Null" and fecha_final != "Null":
                error.append({"field": "Meses,FechaInicio,FechaFinal", "data": "",
                              "message": "No se puede buscar por meses y fechas al mismo tiempo"})

            if meses != "Null":
                if int(meses) > 2:
                    error.append({"field": "Meses", "data": meses,
                                  "message": "El rango de busqueda por mes no puede ser mayor que 2"})
                else:  # -----------> Query con filtro de meses (1 mes o 2 meses)
                    Fecha_actual = datetime.date.today()
                    Fecha_hasta_hora = str(Fecha_actual) + " 23:59:59"
                    fecha_final = datetime.datetime.strptime(Fecha_hasta_hora, "%Y-%m-%d %H:%M:%S")
                    if int(meses) == 1:
                        fecha_inicio = Fecha_actual.replace(day=1)
                    else:
                        fecha_inicio = Fecha_actual - relativedelta(months=int(1))
                    Fecha_inicio_hora = str(fecha_inicio) + " 00:00:00"
                    fecha_inicio = datetime.datetime.strptime(Fecha_inicio_hora, "%Y-%m-%d %H:%M:%S")

            if meses == "Null" and fecha_inicio != "Null" or "" and fecha_final != "Null" or "":  # ----------> Query con filtro de fechas
                FechaDesdeHora = str(fecha_inicio) + " 00:00:00"
                FechaHastaHora = str(fecha_final) + " 23:59:59"
                fecha_inicio = datetime.datetime.strptime(FechaDesdeHora, "%Y-%m-%d %H:%M:%S")
                fecha_final = datetime.datetime.strptime(FechaHastaHora, "%Y-%m-%d %H:%M:%S")

            ingresos = transferencia.objects.filter(cta_beneficiario=request.query_params["NumeroCuenta"],
                                                    fecha_creacion__gte=fecha_inicio,
                                                    fecha_creacion__lte=fecha_final).order_by("-fecha_creacion")

        if error:
            MensajeError(error)
        n = 0
        list = []
        for datos in ingresos:
            tipo = tipo_transferencia.objects.get(id=datos.tipo_pago_id)
            fecha, hora = str(datos.fecha_creacion).split(" ")
            origen = "Individual"
            if datos.masivo_trans_id != None:
                origen = "Masiva"
            diccionario = {
                "id": datos.id,
                "Beneficiario": datos.nombre_beneficiario,
                "Monto": f'${datos.monto}',
                "FechaOperacion": fecha.replace("-", "/ "),
                "Tipo": tipo.nombre_tipo,
                "Origen": origen
            }
            list.append(diccionario)
            n = n + 1  # ---------> Mejora la velocidad que retorna el listado
            if n == size:
                break
        page = self.paginate_queryset(list)
        return Response(page, status=status.HTTP_200_OK)


# (2021-12-20 Jose) End point pata listar/filtro de movimientos (ingresos/egresos)
class MovimientosIngresosEgresos(GenericViewSet):
    permission_classes = ()
    pagination_class = PageNumberPagination

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request):
        error = []
        meses = request.query_params["Meses"]
        fecha_inicio = request.query_params["FechaInicio"]
        fecha_final = request.query_params["FechaFinal"]

        size = NumInt(size=self.request.query_params["size"])
        self.pagination_class.page_size = size

        if meses == "Null" and fecha_inicio == "Null" and fecha_final == "Null":  # ----------> Query sin filtro
            ingresos = transferencia.objects.filter(cta_beneficiario=request.query_params["NumeroCuenta"]).order_by(
                "-fecha_creacion")
            egresos = transferencia.objects.filter(cuenta_emisor=request.query_params["NumeroCuenta"]).order_by(
                "-fecha_creacion")

        else:
            if meses != "Null" and fecha_inicio != "Null" and fecha_final != "Null":
                error.append({"field": "Meses,FechaInicio,FechaFinal", "data": "",
                              "message": "No se puede buscar por meses y fechas al mismo tiempo"})

            if meses != "Null":
                if str(meses) not in [1, 2]:
                    error.append({"field": "Meses", "data": meses,
                                  "message": "El rango de busqueda por mes no puede ser mayor que 2"})
                else:  # -----------> Query con filtro de meses (1 mes o 2 meses)

                    Fecha_actual = datetime.date.today()
                    Fecha_hasta_hora = str(Fecha_actual) + " 23:59:59"
                    fecha_final = datetime.datetime.strptime(Fecha_hasta_hora, "%Y-%m-%d %H:%M:%S")
                    if int(meses) == 1:
                        fecha_inicio = Fecha_actual.replace(day=1)
                    else:
                        fecha_inicio = Fecha_actual.replace(day=1) - relativedelta(months=int(1))
                    Fecha_inicio_hora = str(fecha_inicio) + " 00:00:00"
                    fecha_inicio = datetime.datetime.strptime(Fecha_inicio_hora, "%Y-%m-%d %H:%M:%S")

            if meses == "Null" and fecha_inicio != "Null" or "" and fecha_final != "Null" or "":  # ----------> Query con filtro de fechas
                FechaDesdeHora = str(fecha_inicio) + " 00:00:00"
                FechaHastaHora = str(fecha_final) + " 23:59:59"
                fecha_inicio = datetime.datetime.strptime(FechaDesdeHora, "%Y-%m-%d %H:%M:%S")
                fecha_final = datetime.datetime.strptime(FechaHastaHora, "%Y-%m-%d %H:%M:%S")

            ingresos = transferencia.objects.filter(cta_beneficiario=request.query_params["NumeroCuenta"],
                                                    fecha_creacion__gte=fecha_inicio,
                                                    fecha_creacion__lte=fecha_final).order_by("-fecha_creacion")
            egresos = transferencia.objects.filter(cuenta_emisor=request.query_params["NumeroCuenta"],
                                                   fecha_creacion__gte=fecha_inicio,
                                                   fecha_creacion__lte=fecha_final).order_by("-fecha_creacion")

        if error:
            MensajeError(error)
        n = 0
        list = []
        for datos in egresos:
            tipo = tipo_transferencia.objects.get(id=datos.tipo_pago_id)
            fecha, hora = str(datos.fecha_creacion).split(" ")
            origen = "Individual"
            if datos.masivo_trans_id != None:
                origen = "Masiva"
            diccionario = {
                "id": datos.id,
                "Beneficiario": datos.nombre_beneficiario,
                "Monto": f'-${datos.monto}',
                "FechaOperacion": fecha.replace("-", "/ "),
                "Tipo": tipo.nombre_tipo,
                "Origen": origen
            }
            list.append(diccionario)
            n = n + 1  # ---------> Mejora la velocidad que retorna el listado
            if n == size:
                break

        for datos in ingresos:
            tipo = tipo_transferencia.objects.get(id=datos.tipo_pago_id)
            fecha, hora = str(datos.fecha_creacion).split(" ")
            origen = "Individual"
            if datos.masivo_trans_id != None:
                origen = "Masiva"
            diccionario = {
                "id": datos.id,
                "Beneficiario": datos.nombre_beneficiario,
                "Monto": f'${datos.monto}',
                "FechaOperacion": fecha.replace("-", "/ "),
                "Tipo": tipo.nombre_tipo,
                "Origen": origen
            }
            list.append(diccionario)
            n = n + 1  # ---------> Mejora la velocidad que retorna el listado
            if n == size:
                break

        lista_ordenada = sorted(list, key=itemgetter('id'), reverse=True)
        page = self.paginate_queryset(lista_ordenada)
        return Response(page, status=status.HTTP_200_OK)


class DescargarPdf(GenericViewSet):
    permission_classes = ()

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request):
        # queryset = transferencia.objects.get(id=18963)
        # EstadoCuenta(username)
        # cosas por ordenar
        # masivoId = 767
        # Pido datos
        username_pdf = self.request.query_params['username']
        # pedir fechas
        fecha_inicio = self.request.query_params['FechaInicio']
        fecha_final = self.request.query_params['FechaFinal']
        # perdir numero cuenta
        numero_cuenta = self.request.query_params['NumeroCuenta']
        cuenta_clave = self.request.query_params['CuentaClave']
        # pido username
        username_pdf = self.request.query_params['username']
        # Ordenar las fechas como lo pide el periodo
        fecha_inicio_periodo = fecha_inicio
        fecha_final_periodo = fecha_final
        # ordenar fechas como lo pide la DB
        fecha_inicio = str(fecha_inicio) + ' 00:00:00'
        fecha_final = str(fecha_final) + " 23:59:59"
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, "%Y-%m-%d %H:%M:%S")
        fecha_final = datetime.datetime.strptime(fecha_final, "%Y-%m-%d %H:%M:%S")
        # Obtengo datos domicilio,servicio,usuario,cuenta,

        queryset = transferencia.objects.filter(fecha_creacion__gte=fecha_inicio,
                                                fecha_creacion__lte=fecha_final).filter(
            Q(cuenta_emisor=numero_cuenta) |
            Q(cuenta_emisor=cuenta_clave) |
            Q(cta_beneficiario=numero_cuenta) |
            Q(cta_beneficiario=cuenta_clave)
        ).order_by('id')

        if queryset:
            query = queryset.last()
            print(queryset.first())
            saldo_remanente_anterior = queryset.first()
            print(saldo_remanente_anterior.saldo_remanente)
            saldo_remanente_final = queryset.last()
            print(saldo_remanente_final.saldo_remanente)

            cuentaUser = cuenta.objects.get(Q(cuenta=numero_cuenta) | Q(cuentaclave=numero_cuenta))
            servicio = tipo_transferencia.objects.get(id=query.tipo_pago_id)  #########
            usuario = persona.objects.get(id=cuentaUser.persona_cuenta_id)
            domicilioUser = domicilio.objects.get(domicilioPersona_id=usuario.id)
            # For para sacar datos de la primera hoja
            numero_cargos = 0  # -----> cantidad de numero de cargos realizados
            cantidad_cargos = 0  # -----> total de dinero de cargos
            numero_abonos = 0  # -----> cantidad de numero de abonos realizados
            cantidad_abonos = 0  # -----> total de dinero de abonos
            for datos_primera_hoja in queryset:
                print(datos_primera_hoja.cuenta_emisor)
                if str(datos_primera_hoja.cuenta_emisor) == str(
                        numero_cuenta) or datos_primera_hoja.cuenta_emisor == str(cuenta_clave):
                    print('entra condicion 1')
                    numero_cargos = numero_cargos + 1
                    cantidad_cargos = cantidad_cargos + datos_primera_hoja.monto

                if str(datos_primera_hoja.cta_beneficiario) == str(
                        numero_cuenta) or datos_primera_hoja.cta_beneficiario == str(cuenta_clave):
                    print('entra condicion 2')
                    numero_abonos = numero_abonos + 1
                    cantidad_abonos = cantidad_abonos + datos_primera_hoja.monto
                promedio_saldos_diarios_abonos = 0
                promedio_saldos_diarios_cargos = 0
                if numero_abonos != 0:
                    promedio_saldos_diarios_abonos = float(cantidad_abonos) / float(numero_abonos)
                if numero_cargos != 0:
                    promedio_saldos_diarios_cargos = float(cantidad_cargos) / float(numero_cargos)
                datos = queryset.count()
                numPagTotal = CalcularHojasTotales(datos)
                # Creamos el pdf desde cero
                pdf_file = 'TMP/web/Estado_Cuentas/PDF/Estado-cuenta-' + username_pdf + ".pdf"
                # direccion de las imagenes para el pdf
                carpetaImgEstadoCuenta = "TEMPLATES/web/EstadoCuenta/Img/"
                # Defino el pdf
                can = canvas.Canvas(pdf_file, pagesize=letter)
                # Coloco las imagenes para el pdf
                can.drawImage(carpetaImgEstadoCuenta + "Estado-cuenta.jpg", 0, 660, width=612, height=125)  # Logo 660
                can.drawImage(carpetaImgEstadoCuenta + "Informacion-periodo.jpg", 230, 585, width=233,
                              height=60)  # Informacion del periodo
                can.drawImage(carpetaImgEstadoCuenta + "informacion-financiera-detalles-cargos.jpg", 34, 355, width=545,
                              height=155)  # 355
                can.drawImage(carpetaImgEstadoCuenta + "Detalles-movimientos-realizados.jpg", 33, 273, width=547,
                              height=80)
                can.drawImage(carpetaImgEstadoCuenta + "Pie-pagina.jpg", 33, 25, width=408, height=27)
                # Coloco datos personales que pide el pdf
                can.setFontSize(8.5)
                can.setFillColor("black")
                can.drawString(34, 655, usuario.name)  # empresa
                can.drawString(34, 642, servicio.nombre_tipo)  # nombre del servicio ver bien que hacer con esto
                can.setFillColor(HexColor('#5c5a59'))
                can.drawString(34, 630, "No. Cuenta: " + str(cuentaUser.cuenta))  # NO. cuenta
                can.drawString(34, 617, "No. Cliente / Prefijo: " + str(usuario.id))  # NO.CLIENTE/b PREFiJO
                can.drawString(34, 605, "RFC: " + usuario.rfc)  # RFC
                can.drawString(34, 585, domicilioUser.calle + " " + domicilioUser.no_exterior)  # Calle y numero
                can.drawString(34, 570, domicilioUser.colonia)  # Colonia
                can.drawString(34, 555, domicilioUser.alcaldia_mpio)  # Delegacion
                can.drawString(34, 540, "CP " + domicilioUser.codigopostal + " " + domicilioUser.estado)  # CP y Estado
                # Periodo y fecha de corte
                anio_inicio, mes, dia_inicio = (fecha_inicio_periodo).split('-')
                mes_inicio_periodo = Meses(mes)
                anio_final, mes, dia_final = (fecha_final_periodo).split('-')
                mes_final_periodo = Meses(mes)
                periodo = 'Del ' + anio_inicio + '-' + mes_inicio_periodo + '-' + dia_inicio + ' al ' + anio_final + '-' + mes_final_periodo + '-' + dia_final
                can.drawString(295, 611, periodo)  # Periodo
                can.drawString(295, 592, anio_final + '-' + mes_final_periodo + '-' + dia_final)  # Corte
                # detalle de cargos
                can.drawString(395, 465, "Este producto no genera ninguna comisión")  # commissions Cobradas
                can.drawString(395, 445, "n/a")  # Impuestos retenidos
                can.drawString(395, 425, "n/a")  # cargos objetados
                # informacion financiera
                # -separar saldo final y anterior
                try:
                    saldo = saldo_remanente_anterior.saldo_remanente
                    can.drawRightString(291, 465, f'{"$"}{saldo:3,.2f}')
                except:
                    entero = saldo_remanente_anterior.saldo_remanente
                    if entero != None or entero != "":
                        can.drawRightString(291, 465, "NA")
                    else:
                        can.drawRightString(291, 465, f'{"$"}{entero:3,.2f}')
                saldo = float(cantidad_abonos)
                can.drawRightString(291, 445, f'{"$"}{saldo:3,.2f}')  # Depositos y abonos $
                can.drawCentredString(165, 445, str(numero_abonos))  # Depositos y abonos total

                # - separar saldo y centavos de cargos
                saldo = float(cantidad_cargos)
                can.drawRightString(291, 425, f'{"$"}{saldo:3,.2f}')  # Retiros y cargos $
                can.drawCentredString(165, 425, str(numero_cargos))  # Retiros y cargos total
                try:
                    saldo = saldo_remanente_final.saldo_remanente
                    can.drawRightString(291, 405, f'{"$"}{saldo:3,.2f}')  # Saldo final
                except:
                    entero = saldo_remanente_final.saldo_remanente
                    try:
                        can.drawRightString(291, 405, f'{"$"}{entero:3,.2f}')
                    except:
                        can.drawRightString(291, 405, "NA")

                # -separar saldo y centavos de promedio diarios
                saldo = float(promedio_saldos_diarios_abonos)
                can.drawRightString(291, 385, f'{"$"}{saldo:3,.2f}')  # Promedio saldo diarios abonos
                saldo = float(promedio_saldos_diarios_cargos)
                can.drawRightString(291, 365, f'{"$"}{saldo:3,.2f}')  # Promedio saldos diarios cargos

                # Detalle de movimientos realizados
                can.setFontSize(8.4)
                pagina = 1
                can.setFillColor("gray")
                can.setFontSize(9.1)
                can.drawRightString(580, 30, str(pagina) + "/" + str(numPagTotal))  # ------
                can.setFontSize(8.4)
                can.setFillColor(HexColor('#5c5a59'))

                # Calculo de lineas
                can.setLineWidth(0)
                lineY = 276
                lineYD = 295  # --------> coordenadad eje y pada los datos
                lineXFO = 55  # -------> fecha operacion 80
                lineXB = 130  # --------> fecha liquidacion 218
                lineXN = 190  # ------------> descripcion y clave
                lineXM = 409  # ------------> cargos de la operacion
                lineXMA = 475  # ------------> abonos de la operacion
                lineXS = 545  # -----------> saldo anterior

                # locale.setlocale(locale.LC_ALL, 'en_US') #################################
                for datos in queryset:
                    # Nueva pagina
                    if lineY < 66 and lineYD < 80:
                        # Colocar imagenes a la nueva pagina y se cambian las coordenadas de las lineas 
                        can.showPage()
                        can.setLineWidth(0)
                        lineY = 696
                        lineYD = 715
                        can.drawImage(carpetaImgEstadoCuenta + "Detalles-movimientos-realizados.jpg", 33, 694,
                                      width=547, height=80)
                        can.drawImage(carpetaImgEstadoCuenta + "Pie-pagina.jpg", 33, 25, width=408, height=27)
                        can.setFillColor("gray")
                        can.setFontSize(9.1)
                        pagina = int(pagina) + 1
                        can.drawRightString(580, 30, str(pagina) + "/" + str(numPagTotal))
                    can.setLineWidth(0)
                    can.setFontSize(7.5)
                    can.setFillColor(HexColor('#5c5a59'))
                    # Fecha operacion
                    fecha_operacion, tiempo = str(datos.fecha_creacion).replace('-', '/').split(' ')
                    can.drawCentredString(lineXFO, lineYD, fecha_operacion)
                    # Fecha liquidacion
                    try:
                        fecha_liquidacion, segundos = str(datos.date_modify).replace("-", "/").split(".")
                        tiempo = "am"
                        fecha_liquidacion_real, hora = fecha_liquidacion.replace(":", "").split(" ")
                        if int(hora) > 120000:
                            tiempo = "pm"
                        can.drawCentredString(lineXB, lineYD, fecha_liquidacion + ' ' + tiempo)
                    except:
                        can.drawCentredString(lineXB, lineYD, "NA")
                    # Descripcion
                    can.drawString(lineXN, lineYD, datos.concepto_pago)
                    # CVE rastreo
                    lineYDN = lineYD - 13
                    can.drawString(lineXN, lineYDN, 'CVE RASTREO: ' + datos.clave_rastreo)
                    if str(datos.cuenta_emisor) == str(numero_cuenta) or str(datos.cuenta_emisor) == str(cuenta_clave):
                        Monto = datos.monto
                        can.drawCentredString(lineXM, lineYD, f'{"$"}{Monto:3,.2f}')
                    else:
                        Monto = datos.monto
                        can.drawCentredString(lineXMA, lineYD, f'{"$"}{Monto:3,.2f}')
                    # saldo remanente
                    try:
                        Monto = datos.saldo_remanente
                        can.drawCentredString(lineXS, lineYD, f'{"$"}{Monto:3,.2f}')
                    except:
                        Monto = '0', '00'
                    # can.drawCentredString(lineXS, lineYD, f'{"$"}{Monto:3,.2f}')
                    can.setLineWidth(0)
                    can.line(34, lineY, 578, lineY)
                    lineY = lineY - 30
                    lineYD = lineYD - 30
                can.save()
            filename = 'TMP/web/Estado_Cuentas/PDF/Estado-cuenta-' + username_pdf + ".pdf"
            filepath = filename
            path = open(filepath, 'r')
            mime_type, _ = mimetypes.guess_type(filepath)
            response = FileResponse(open(filename, 'rb'))
            response['Content-Disposition'] = "attachment; filename=%s" % filename
            return response
        else:
            error = {'field': '', "data": '', 'message': "No se encontro nigun registro"}
            MensajeError(error)


# (Jose 2021/12/28) Crear y retornar Excel de su estado de cuenta

class GenerarExcel(GenericViewSet):
    permission_classes = ()

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request):
        # Pedif fechas y numnero de cuenta
        fecha_inicio = self.request.query_params['FechaInicio']
        fecha_final = self.request.query_params['FechaFinal']
        numero_cuenta = self.request.query_params['NumeroCuenta']
        username_excel = self.request.query_params['username']
        # Ordenar fecha como lo pide el Excel
        fecha_inicio_periodo = fecha_inicio
        fecha_final_periodo = fecha_final
        # Ordenar fecha como lo pide la DB
        fecha_inicio = str(fecha_inicio) + ' 00:00:00'
        fecha_final = str(fecha_final) + " 23:59:59"
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, "%Y-%m-%d %H:%M:%S")
        fecha_final = datetime.datetime.strptime(fecha_final, "%Y-%m-%d %H:%M:%S")
        # hacer un query en trasacciones de egresos e ingresos y ordenar por id
        registros = transferencia.objects.filter(
            Q(cuenta_emisor=numero_cuenta, fecha_creacion__gte=fecha_inicio, fecha_creacion__lte=fecha_final) | Q(
                cta_beneficiario=numero_cuenta, fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_final)).order_by('id')
        # Texto final
        texto_final = 'Los recursos de los Usuarios en las operaciones realizadas con Polipay  no se encuentran garantizados por ninguna autoridad. Los fondos de pago electrónico no generan rendimientos o beneficios monetarios por los saldos acumulados en los mismos. Polipay  recibe consultas, reclamaciones o aclaraciones, en su Unidad Especializada de Atención a Usuarios, por correo electrónico a contacto@polipay.com . En el caso de no obtener una respuesta satisfactoria, podrá acudir a la Comisión Nacional para la Protección y Defensa de los Usuarios de Servicios Financieros a través de su página web: https//gob.mx/condusef o al número telefónico 5553400999. '
        # Abri el excel
        excel_estado_cuenta = load_workbook(filename="TEMPLATES/web/EstadoCuenta/Excel/Estado-Cuenta.xlsx")
        sheet = excel_estado_cuenta.active
        # Obtener datos del usuario
        datos_cuenta = cuenta.objects.get(cuenta=numero_cuenta)
        datos_user = persona.objects.get(id=datos_cuenta.persona_cuenta_id)
        direccion = domicilio.objects.get(domicilioPersona_id=datos_user.id)
        # Periodo
        periodo = 'Del ' + str(fecha_inicio_periodo) + ' al ' + str(fecha_final_periodo)
        sheet['G18'] = periodo
        # Colocar datos del usuario en el excel
        sheet['B11'] = datos_user.name
        sheet['B11'].font = Font(u'Arial', bold=True, size=11)
        sheet['C13'] = datos_cuenta.cuenta
        sheet['C14'] = datos_user.id
        sheet['F11'] = datos_user.name
        domicilio_user = direccion.calle + ' ' + direccion.no_exterior + ', COL ' + direccion.colonia \
                         + ', DELEG ' + direccion.alcaldia_mpio + ', CP ' + direccion.codigopostal + ', ' + direccion.estado
        sheet['F12'] = domicilio_user
        numero = 31

        numero_abonos_depositos = 0
        total_abonos_depositos = 0
        numero_retiros_cargos = 0
        total_retiros_cargos = 0
        saldo_anterior = registros.first()
        saldo_final = registros.last()
        # Colocamos saldo anterior y saldo final
        sheet['D20'] = saldo_anterior.saldo_remanente
        sheet['D23'] = saldo_final.saldo_remanente
        # B ---> Fecha de operacion
        # C ---> Fecha de liquidacion
        # D ---> concepto
        # E ---> Clave de rastreo
        # F ---> Cargos
        # G ---> Abonos
        # H ---> Saldo
        # hacer un for de la trasferencia
        for registro in registros:
            # Acomodar la fecha para el excel
            fecha_creacion, tiempo = str(registro.fecha_creacion).split(' ')
            # Acomodar los datos en las celdas
            sheet["B" + str(numero)] = fecha_creacion
            sheet['B' + str(numero)].alignment = Alignment(horizontal="left")
            sheet["C" + str(numero)] = registro.date_modify
            sheet['C' + str(numero)].alignment = Alignment(horizontal="left")
            sheet["D" + str(numero)] = registro.concepto_pago
            sheet['D' + str(numero)].alignment = Alignment(horizontal="left")
            sheet["E" + str(numero)] = registro.clave_rastreo
            sheet['E' + str(numero)].alignment = Alignment(horizontal="left")
            # Ordenar el monto con centavos
            # locale.setlocale(locale.LC_ALL, 'en_US') @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
            # if '.' in str(registro.monto):
            #     print(registro.id)
            #     print('tiene punto decimal')
            saldo, centavos = str(registro.monto).split(".")
            # else:
            #     print('no tiene punto decimal')
            #     saldo = registro.monto
            #     centavos = '00'
            # Ver si es egreso o ingreso y acomodarlo 
            if numero_cuenta == registro.cuenta_emisor:
                sheet["F" + str(numero)] = "$" + str(f'{int(saldo):n}') + '.' + str(centavos[:2])
                sheet['F' + str(numero)].alignment = Alignment(horizontal="left")
                numero_retiros_cargos = numero_retiros_cargos + 1
                total_retiros_cargos = float(total_retiros_cargos) + float(registro.monto)
            else:
                sheet["G" + str(numero)] = "$" + str(f'{int(saldo):n}') + '.' + str(centavos[:2])
                sheet['G' + str(numero)].alignment = Alignment(horizontal="left")
                numero_abonos_depositos = numero_abonos_depositos + 1
                total_abonos_depositos = float(total_abonos_depositos) + float(registro.monto)
            # Ordenar saldo
            # if registro.saldo_remanente is '.':
            #     saldo, centavos = str(registro.saldo_remanente).split(".")
            # else:
            #     saldo = registro.saldo_remanente
            #     centavos = '00'
            try:
                saldo_remanente = float(registro.saldo_remanente)
                saldo, centavos = str(saldo_remanente).split(".")
            except:
                saldo, centavos = '0', '000'
            try:
                sheet["H" + str(numero)] = '$' + str(f'{int(saldo):n}') + '.' + str(centavos[:2])
                sheet['H' + str(numero)].alignment = Alignment(horizontal="left")
            except:
                sheet["H" + str(numero)] = 'NA'
                sheet['H' + str(numero)].alignment = Alignment(horizontal="left")
            numero = int(numero) + 1
        numero_inicio = numero + 2
        numero_final = numero + 4
        sheet.merge_cells('B' + str(numero_inicio) + ':H' + str(numero_final))
        cell = sheet.cell(row=numero_inicio, column=2)
        cell.value = texto_final
        cell.alignment = cell.alignment.copy(wrapText=True)  # ----> Sirve para ajustar el texto en las celdas
        # Colocamos cantidad de abonos/depositos y retiros/cargos
        sheet['C21'] = numero_abonos_depositos
        sheet['C22'] = numero_retiros_cargos
        # Colocamos total de depositos y cargos
        sheet['D21'] = total_abonos_depositos
        sheet['D22'] = total_retiros_cargos
        fecha_actual = str(datetime.date.today()).replace('-', '')
        excel_estado_cuenta.save(
            filename='TMP/web/Estado_Cuentas/Excel/Estado-cuenta-' + username_excel + '_' + fecha_actual + ".xlsx")
        # Descargar Excel
        filename = 'TMP/web/Estado_Cuentas/Excel/Estado-cuenta-' + username_excel + '_' + fecha_actual + ".xlsx"
        filepath = filename
        path = open(filepath, 'r')
        mime_type, _ = mimetypes.guess_type(filepath)
        response = FileResponse(open(filename, 'rb'))
        response['Content-Disposition'] = "attachment; filename=%s" % filename
        return response
