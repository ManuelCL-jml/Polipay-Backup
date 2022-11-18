import datetime as dt
import random
import mimetypes
import locale
from dataclasses import dataclass
from typing import Optional, ClassVar, NoReturn

from django.core.exceptions import ObjectDoesNotExist
from django.utils.datastructures import MultiValueDictKeyError
from django.db import IntegrityError
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page

from rest_framework.pagination import PageNumberPagination
from rest_framework.viewsets import GenericViewSet
from rest_framework.generics import ListAPIView, RetrieveAPIView
from rest_framework.response import Response
from rest_framework import status

from django.http.response import FileResponse

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from MANAGEMENT.ComissionPay.comission import RegistraOrdenDispersionMasivaIndividual
from MANAGEMENT.Standard.errors_responses import MyHttpError
from MANAGEMENT.Standard.success_responses import MyHtppSuccess
from MANAGEMENT.Utils.utils import remove_asterisk
from apps.api_stp.exc import StpmexException
from apps.commissions.models import Commission_detail
from apps.logspolipay.manager import RegisterLog
from apps.permision.permisions import BlocklistPermissionV2
from apps.transaction.exc import ParamsNotProvided
from apps.transaction.messages import send_massive_email
from apps.transaction.api.web.serializers.serializers_dispersiones import *
from apps.users.models import cuenta, persona, tarjeta, domicilio
from apps.transaction.models import *
from apps.users.management import filter_data_or_return_none, get_id_cuenta_eje, \
    get_data_empresa
from polipaynewConfig.settings import COST_CENTER_POLIPAY_COMISSION, COST_CENTER_INNTEC


def create_folio():
    return random.randrange(100000, 999999, 6)


# (ChrGil 01.10.2021) Clase que permite verificar el monto actual de un saldo de una cuenta.
class VerifyMonto(RetrieveAPIView):
    @method_decorator(cache_page(60 * 0.1))
    def retrieve(self, request, *args, **kwargs):
        razon_social_id: int = self.request.query_params['razon_social_id']
        get_monto = cuenta.objects.filter(persona_cuenta_id=razon_social_id).values('monto').first()
        return Response(get_monto, status=status.HTTP_200_OK)


# (ChrGil 15.09.2021) Clase que regresa un listado de personal externo
# Endpoint: http://127.0.0.1:8000/transaction/admin/v2/RecNomBen/list/
class RecomendarNombreBeneficiario(ListAPIView):
    def get_queryset(self, *args, **kwargs) -> List:
        queryset: List = grupoPersona.objects.filter(**kwargs).values('person_id')
        list_data: List = []
        for query in queryset:
            data = cuenta.objects.filter(
                persona_cuenta_id=query['person_id'],
                is_active=True
            ).values('id', 'persona_cuenta__name', 'persona_cuenta__last_name', 'persona_cuenta__email')

            if data:
                list_data.append(data[0])

        return list_data

    def list(self, request, *args, **kwargs):
        cuenta_eje: int = get_id_cuenta_eje(request.user.get_only_id())
        lista = self.get_queryset(empresa_id=cuenta_eje, relacion_grupo_id=6)
        return Response(lista)


# # (ChrGil 15.09.2021) Clase que regresa un listado de personal externo
# # Endpoint: http://127.0.0.1:8000/transaction/admin/v2/RecNomBen/list/
# class RecomendarNombreBeneficiario(ListAPIView):
#     def get_queryset(self, *args, **kwargs) -> List:
#         queryset: List = grupoPersona.objects.filter(**kwargs).values('person_id')
#         list_data: List = []
#         for query in queryset:
#             data = cuenta.objects.filter(
#                 persona_cuenta_id=query['person_id'],
#                 is_active=True
#             ).values('id', 'persona_cuenta__name', 'persona_cuenta__last_name', 'persona_cuenta__email')
#
#             if data:
#                 list_data.append(data[0])
#
#         return list_data
#
#     def list(self, request, *args, **kwargs):
#         cuenta_eje: int = get_id_cuenta_eje(request.user.get_only_id())
#         lista = self.get_queryset(empresa_id=cuenta_eje, relacion_grupo_id=6)
#         return Response(lista)


@dataclass
class CreateDispersionMasiva:
    observation: str
    type_dispersion: str
    serializer_class = SerializerDisMassivas

    def create(self) -> int:
        data: Dict = {"observations": self.observation}
        serializer = self.serializer_class(data=data)
        serializer.is_valid(raise_exception=True)
        dispersion_id = serializer.create(serializer.data)
        return dispersion_id

    def is_massive(self):
        if self.type_dispersion.upper() == 'M':
            return self.create()
        return None


@dataclass
class CreateDispersionV2:
    person_list: List[Dict]
    context: Dict[str, Any]
    serializer_class = SerializerDispersionTest

    def send_massive_email(self, list_beneficiario: List[Dict], list_emisor: List[Dict]) -> bool:
        return send_massive_email(list_beneficiario, list_emisor)

    def create_dispersion(self, list_validation: List, serializer: Optional) -> bool:
        try:
            with atomic():
                for validated_data in list_validation:
                    monto_actual = cuenta.objects.get(cuenta=validated_data['cuenta_emisor']).get_monto_emisor()
                    serializer.create_disper(validated_data, monto_actual)
        except Exception as e:
            err = MyHttpError(
                message="Ocurrio un error inesperado durante el proceso de creación de una dispersión",
                real_error=str(e))
            raise ValidationError(err.standard_error_responses())

        return True

    def validate_all_dispersion(self):
        validation_list: List = []
        list_data_beneficiario: List[Dict] = []
        list_data_emisor: List[Dict] = []
        serializer: Optional = None

        list_data_emisor.append(self.emisor_dict_data(self.context))
        for person in self.person_list:
            serializer = self.serializer_class(data=person, context=self.context)
            serializer.is_valid(raise_exception=True)
            validation_list.append(serializer.data)
            list_data_beneficiario.append(self.beneficiario_dict_data(person, self.context))

            if self.context['type_dispersion'] == "I":
                list_data_emisor[0]['nombre_beneficiario'] = person['nombre_beneficiario']

        self.create_dispersion(validation_list, serializer)
        self.send_massive_email(list_data_beneficiario, list_data_emisor)
        return True

    def emisor_dict_data(self, context: Dict) -> Dict:
        return {
            "folio": create_folio(),
            "email": context['email_admin'],
            "observation": context['observation'],
            "nombre_emisor": context['nombre_emisor'],
            "fecha_operacion": datetime.datetime.now(),
            "monto_total": context['monto_total'],
            "nombre_grupo": context['nombre_grupo']
        }

    def beneficiario_dict_data(self, person: Dict, context: Dict) -> Dict:
        return {
            "folio": create_folio(),
            "name": person['nombre_beneficiario'],
            "email": person['email'],
            "monto": person['monto'],
            "observation": context['observation'],
            "nombre_emisor": context['nombre_emisor'],
            "fecha_operacion": datetime.datetime.now()
        }


# __POSIBLE__OBSOLETO
# (ChrGil 2021-10-22) Creación de una dispersión masiva y individual
class DispersionV2(GenericViewSet):
    permission_classes = (BlocklistPermissionV2,)
    permisos = ["Crear dispersión individual", "Crear dispersión masiva"]

    def create(self, request):
        admin_instance = request.user
        email_admin = admin_instance.get_email()
        get_empresa: Dict = get_data_empresa(admin_instance.get_only_id())
        cuenta_emisor: Dict = cuenta.objects.get(persona_cuenta_id=get_empresa['id']).get_all_cuentas()
        instance_emisor = cuenta.objects.get(persona_cuenta_id=get_empresa['id'])

        for request in request.data:
            observation: str = request['observations']
            person_list: List[Dict] = request.pop('PersonList')
            type_dispersion: str = request['TypeDispersion']
            monto_total: float = request["MontoTotal"]
            nombre_grupo: str = request["NombreGrupo"]

            # masivo_trans: Optional[None, int] = CreateDispersionMasiva(observation, type_dispersion).is_massive()

            context = {
                "empresa": get_empresa['name'],
                "type_dispersion": type_dispersion.upper(),
                "cuenta_emisor": cuenta_emisor,
                "instance_cuenta_emisor": instance_emisor,
                "emisor_id": instance_emisor.id,
                "observation": observation.capitalize(),
                "is_schedule": request['is_schedule'],
                "monto_total": monto_total,
                "nombre_emisor": get_empresa['name'],
                "masivo_trans_id": CreateDispersionMasiva(observation, type_dispersion).is_massive(),
                "logitud_lista": len(person_list),
                "email_admin": email_admin,
                "nombre_grupo": nombre_grupo.upper() if nombre_grupo else None
            }

            CreateDispersionV2(person_list, context).validate_all_dispersion()

        success = MyHtppSuccess('Tu operación se realizo de manera satisfactoria.')
        return Response(success.created(), status=status.HTTP_201_CREATED)


# __POSIBLE__OBSOLETO
# (ChrGil 2021-10-22) Crea una Dispersion masiva, individual ademas envia correos al beneficiario y al emisor
class Dispersion(GenericViewSet):
    serializer_class = SerializerDispersionTest
    serializer_class_massive = SerializerDisMassivas

    def emisor_dict_data(self, context: Dict) -> Dict:
        return {
            "folio": create_folio(),
            "email": context['email_admin'],
            "observation": context['observation'],
            "nombre_emisor": context['nombre_emisor'],
            "fecha_operacion": datetime.datetime.now(),
            "monto_total": context['monto_total'],
            "nombre_grupo": context['nombre_grupo']
        }

    def beneficiario_dict_data(self, person: Dict, context: Dict) -> Dict:
        return {
            "folio": create_folio(),
            "name": person['nombre_beneficiario'],
            "email": person['email'],
            "monto": person['monto'],
            "observation": context['observation'],
            "nombre_emisor": context['nombre_emisor'],
            "fecha_operacion": datetime.datetime.now()
        }

    def send_massive_email(self, list_beneficiario: List[Dict], list_emisor: List[Dict]) -> bool:
        return send_massive_email(list_beneficiario, list_emisor)

    def create_dispersion(self, list_validation: List, serializer: Optional) -> bool:
        for validated_data in list_validation:
            serializer.create(validated_data)
        return True

    def validate_dispersion(self, person_list: List[Dict], context: Dict) -> bool:
        validation_list: List = []
        list_data_beneficiario: List[Dict] = []
        list_data_emisor: List[Dict] = []
        serializer: Optional = None

        list_data_emisor.append(self.emisor_dict_data(context))
        for person in person_list:
            serializer = self.serializer_class(data=person, context=context)
            serializer.is_valid(raise_exception=True)
            validation_list.append(serializer.data)
            list_data_beneficiario.append(self.beneficiario_dict_data(person, context))

            if context['type_dispersion'] == "I":
                list_data_emisor[0]['nombre_beneficiario'] = person['nombre_beneficiario']

        self.create_dispersion(validation_list, serializer)
        self.send_massive_email(list_data_beneficiario, list_data_emisor)
        return True

    def create_dispersion_massive(self, observation: str) -> int:
        data = {'observations': observation}
        serializer_massive = self.serializer_class_massive(data=data)
        serializer_massive.is_valid(raise_exception=True)
        return serializer_massive.create(serializer_massive.data)

    def is_dispersion_massive(self, type_dispersion: str, observation: str):
        if type_dispersion.upper() == 'M':
            return self.create_dispersion_massive(observation)
        return None

    def create(self, request):
        log = RegisterLog(request.user, request)
        admin_instance = request.user
        email_admin = admin_instance.get_email()
        get_empresa: Dict = get_data_empresa(admin_instance.get_only_id())
        cuenta_emisor: Dict = cuenta.objects.get(persona_cuenta_id=get_empresa['id']).get_all_cuentas()
        instance_emisor = cuenta.objects.get(persona_cuenta_id=get_empresa['id'])
        log.json_request(request.data)

        for request in request.data:
            observation: str = request['observations']
            person_list: List[Dict] = request.pop('PersonList')
            type_dispersion: str = request['TypeDispersion']
            monto_total: float = request["MontoTotal"]
            nombre_grupo: str = request["NombreGrupo"]

            masivo_trans: Optional[None, int] = self.is_dispersion_massive(type_dispersion, observation)

            context = {
                "empresa": get_empresa['name'],
                "type_dispersion": type_dispersion.upper(),
                "cuenta_emisor": cuenta_emisor,
                "instance_cuenta_emisor": instance_emisor,
                "observation": observation.capitalize(),
                "is_schedule": request['is_schedule'],
                "monto_total": monto_total,
                "nombre_emisor": admin_instance.get_full_name(),
                "masivo_trans_id": masivo_trans,
                "logitud_lista": len(person_list),
                "email_admin": email_admin,
                "nombre_grupo": nombre_grupo.upper() if nombre_grupo else None
            }

            self.validate_dispersion(person_list, context)
            success = MyHtppSuccess('Tu operación se realizo de manera satisfactoria.', "201", "CREATED")
            log.json_response(success.standard_success_responses())
            return Response(success.standard_success_responses(), status=status.HTTP_201_CREATED)


# Endpoint: http://127.0.0.1:8000/transaction/web/v2/LisStaDis/list/?size=100&date_1=null&date_2=null&nombre_beneficiario=null&nombre_emisor=null
# class ListStatusDispersion(ListAPIView):
#     """
#     Autor: Christian Gil
#     Descripción: Regresa Filtrado de dispersiones individuales, por rango de fechas, nombre beneficiario y nombre emisor
#     Estados:
#         1: Liquidada
#         2: Incompleta
#         3: Pendiente
#         5: Cancelada
#         6: Creada
#     """
#
#     permission_classes = (BlocklistPermissionV2,)
#     permisos = ["Ver dispersiones individuales pendientes", "Ver dispersiones individuales canceladas",
#                 "Ver dispersiones individuales liquidadas", "Ver dispersiones individuales incompletas"]
#     serializer_class = SerializerListDispersiones
#     pagination_class = PageNumberPagination
#
#     def list(self, request, *args, **kwargs):
#         size = self.request.query_params['size']
#         date_1 = request.query_params['date_1']
#         date_2 = request.query_params['date_2']
#         nombre_beneficiario: str = request.query_params['nombre_beneficiario']
#         nombre_emisor: str = request.query_params['nombre_emisor']
#
#         pagination.PageNumberPagination.page_size = size
#         get_empresa: Dict = get_data_empresa(request.user.get_only_id())
#         cuenta_id: int = get_Object_orList_error(cuenta, persona_cuenta_id=get_empresa['id']).get_only_id()
#
#         data = {
#             "tipo_pago_id": 4,
#             "status_trans_id": request.query_params['status_id'],
#             "nombre_beneficiario": '' if nombre_beneficiario == 'null' else nombre_beneficiario,
#             "nombre_emisor": '' if nombre_emisor == 'null' else nombre_emisor,
#             "date_1": datetime.datetime(2000, 12, 31, 00, 00, 00) if date_1 == 'null' else date_1,
#             "date_2": datetime.datetime.now() if date_2 == 'null' else date_2,
#             "cuentatransferencia_id": cuenta_id
#         }
#
#         serializer = self.serializer_class(data=data)
#         serializer.is_valid(raise_exception=True)
#         queryset = serializer.queryset(serializer.data)
#
#         page = self.paginate_queryset(queryset)
#         return self.get_paginated_response(page)


"""
    Filtro: status_id, date_1, date_2, nombre_beneficiario, nombre_emisor, cuenta id (cuenta eje o centro de costos)
    Depende del servicio

"""


class GetInfoCompany:
    info_account: ClassVar[Dict[str, Any]]

    def __init__(self, admin: persona, **kwargs):
        self._admin = admin
        self._razon_social_id = kwargs.get('razon_social_id', None)
        self._get_info_account_cuenta_eje()

        if not self._razon_social_id:
            raise ParamsNotProvided('Operación prohibida, debes de enviar el id del una persona moral')

        if self.info_account.get('rel_cuenta_prod_id') == 3:
            if not self._exists_cost_center:
                raise ValueError('Centro de costos no valido o no existe')

            self._get_info_account_cost_center()

    @property
    def _get_cuenta_eje(self) -> int:
        return get_id_cuenta_eje(self._admin.get_only_id())

    def _get_info_account_cuenta_eje(self):
        self.info_account = cuenta.objects.filter(
            persona_cuenta_id=self._get_cuenta_eje,
            is_active=True
        ).values('id', 'rel_cuenta_prod_id', 'persona_cuenta_id').first()

    @property
    def _exists_cost_center(self) -> bool:
        return grupoPersona.objects.select_related('empresa', 'person', 'relacion_grupo').filter(
            empresa_id=self.info_account.get('persona_cuenta_id'),
            person_id=self._razon_social_id,
            relacion_grupo_id=5
        ).exists()

    def _get_info_account_cost_center(self):
        self.info_account = cuenta.objects.select_related('persona_cuenta').filter(
            persona_cuenta_id=self._razon_social_id,
            is_active=True
        ).values('id', 'cuenta').first()


class ClassListDispersionStatus:
    default_size: ClassVar[int] = 5
    list_dispersion: ClassVar[List[Dict[str, Any]]]

    def __init__(self, razon_social: GetInfoCompany, **kwargs):
        self._status_trans_id = kwargs.get('status_id', None)
        self._nombre_beneficiario = kwargs.get('nombre_beneficiario', '')
        self._nombre_emisor = kwargs.get('nombre_emisor', '')
        self._date_1 = kwargs.get('date_1', dt.date.today() - dt.timedelta(days=91))
        self._date_2 = kwargs.get('date_2', dt.date.today())
        self.size = kwargs.get('size', self.default_size)
        self._cuentatransferencia_id = razon_social.info_account.get('id')
        self._raise_params()
        self.list_dispersion = self.get_is_sheduled

    def _raise_params(self) -> NoReturn:
        if not self._status_trans_id:
            raise ParamsNotProvided('Operación prohibida, debes de enviar el estado de la dispersión')

    @property
    def get_is_sheduled(self) -> List[Dict]:
        list_transaction = self._list

        for i in list_transaction:
            if i.get('programada'):
                shedule = transferenciaProg.objects.filter(transferReferida_id=i.get('id')).values(
                    'fechaProgramada').first()
                i['date_sheduled'] = shedule
        return list_transaction

    @property
    def _list(self) -> List[Dict[str, Any]]:
        return transferencia.objects.select_related(
            'masivo_trans',
            'tipo_pago',
            'status_trans'
        ).filter(
            Q(fecha_creacion__date__gte=self._date_1) &
            Q(fecha_creacion__date__lte=self._date_2)
        ).filter(
            cuentatransferencia_id=self._cuentatransferencia_id,
            status_trans_id=self._status_trans_id,
            nombre_beneficiario__icontains=self._nombre_beneficiario,
            nombre_emisor__icontains=self._nombre_emisor,
            tipo_pago_id=4,
            masivo_trans_id__isnull=True
        ).values(
            'id',
            'nombre_beneficiario',
            'monto',
            'fecha_creacion',
            'date_modify',
            'nombre_emisor',
            'masivo_trans',
            'programada',
            'clave_rastreo'
        ).order_by('-fecha_creacion')


# (ChrGil 2022-02-12) Listar dispersiones individuales por estado
class ListStatusDispersion(ListAPIView):
    permisos = ["Ver dispersiones individuales pendientes", "Ver dispersiones individuales canceladas",
                "Ver dispersiones individuales liquidadas", "Ver dispersiones individuales incompletas"]

    permission_classes = (BlocklistPermissionV2,)
    pagination_class = PageNumberPagination

    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        try:
            admin: persona = request.user
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            log.json_request(request.query_params)
            emisor = GetInfoCompany(admin, **data)
            dispersion_list = ClassListDispersionStatus(emisor, **data)
            self.pagination_class.page_size = dispersion_list.size
            return self.get_paginated_response(self.paginate_queryset(dispersion_list.list_dispersion))
        except (ObjectDoesNotExist, ValueError) as e:
            err = MyHttpError('Su cuenta no esta asociada a una persona moral', real_error=str(e))
            log.json_response(err.standard_error_responses())
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)


class DetailDispercionIndividual(GenericViewSet):
    permission_classes = ()
    serializer_class = serializerDetailDispercionIndividual

    def get_queryset(self, *args, **kwargs):
        return filter_data_or_return_none(*args, **kwargs)

    def list(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        person_instance = persona.objects.get(id=self.request.query_params['usuario'])
        id = self.request.query_params['dispersion']
        log.json_request(request.query_params)
        account_instance = cuenta.objects.get(persona_cuenta_id=person_instance.id)
        queryset = transferencia.objects.filter(id=id, cuentatransferencia_id=account_instance.id, tipo_pago_id=4)

        context = {
            "instance_cta_benef": queryset
        }

        serializer = self.serializer_class(instance=queryset, many=True, context=context)
        return Response(serializer.data, status=status.HTTP_200_OK)


# class ListDispersionesMasivasEstado(ListAPIView):
#     """
#     Autor: Christian Gil
#     Descripción: Regresa listado de estados de dispersion masiva con filtrado por
#                  rango de fechas, por estado y por concepto pago
#
#     """
#     permission_classes = (BlocklistPermissionV2,)
#     permisos = ["Ver dispersiones masivas pendientes", "Ver dispersiones masivas canceladas",
#                 "Ver dispersiones masivas procesadas"]
#     pagination_class = PageNumberPagination
#
#     def get_queryset(self, *args, **kwargs):
#         return filter_data_or_return_none(*args, **kwargs)
#
#     def data_filter(self, observation: str, date_1: datetime, date_2: datetime) -> Dict:
#         return {
#             "observations": '' if observation == 'null' else observation,
#             "date_1": datetime.datetime(2000, 12, 31, 00, 00, 00) if date_1 == 'null' else date_1,
#             "date_2": datetime.datetime.now() if date_2 == 'null' else date_2,
#         }
#
#     def conjunto(self, querysets: List) -> List:
#         conjunto_ids = set()
#         for index in querysets:
#             conjunto_ids.add(index['masivo_trans_id'])
#
#         return list(conjunto_ids)
#
#     def list_dispersiones(self, list_masivo_trans_id: List) -> List:
#         lista = [
#             transferencia.objects.filter(
#                 masivo_trans_id=masivo_trans_id).values(
#                 'masivo_trans_id',
#                 'nombre_emisor',
#                 'masivo_trans__observations',
#                 'masivo_trans__date_liberation').first()
#
#             for masivo_trans_id in list_masivo_trans_id
#         ]
#         return lista
#
#     def list(self, request, *args, **kwargs):
#         PageNumberPagination.page_size = self.request.query_params['size']
#         status_trans_id: int = request.query_params['status_id']
#         observations: str = request.query_params['observations']
#         date_1 = request.query_params['date_1']
#         date_2 = request.query_params['date_2']
#
#         cuenta_eje: int = get_id_cuenta_eje(request.user.get_only_id())
#         cuenta_id: int = cuenta.objects.get(persona_cuenta_id=cuenta_eje).get_only_id()
#
#         data_filter: Dict = self.data_filter(observations, date_1, date_2)
#         querysets = transferencia.objects.all().filter(
#             cuentatransferencia_id=cuenta_id,
#             status_trans_id=status_trans_id,
#             masivo_trans_id__isnull=False,
#             masivo_trans__observations__icontains=data_filter['observations'],
#             masivo_trans__date_liberation__range=(data_filter['date_1'], data_filter['date_2']),
#         ).values('masivo_trans_id')
#
#         list_masivo_trans_id: List = self.conjunto(querysets)
#         ordenamiento_desc: List = sorted(list_masivo_trans_id, reverse=True)
#         data_list: List = self.list_dispersiones(ordenamiento_desc)
#
#         page = self.paginate_queryset(data_list)
#         return self.get_paginated_response(page)

# size, status_id, observations, date_1, date_2, request.user


class ClassListDispersionMassive:
    defaul_size: ClassVar[int] = 5
    dispersion_massive: ClassVar[List[Dict[str, Any]]]

    def __init__(self, emisor: GetInfoCompany, **kwargs):
        self._status_id = kwargs.get('status_id', None)
        self._observations = kwargs.get('observations', '')
        self._start_date = kwargs.get('date_1', dt.date.today() - dt.timedelta(days=91))
        self._end_date = kwargs.get('date_2', dt.date.today())
        self.size = kwargs.get('size', self.defaul_size)
        self._raise_params()
        self._emisor = emisor
        self.dispersion_massive = [self._render_json_transaction(**trans) for trans in self._list]

    def _raise_params(self) -> NoReturn:
        if not self._status_id:
            raise ParamsNotProvided('Operación prohibida, debes de enviar el estado de la dispersión')

    @property
    def _list_id_massive(self) -> List[int]:
        massive = transferencia.objects.select_related(
            'masivo_trans',
            'tipo_pago',
            'status_trans',
            'cuentatransferencia'
        ).filter(
            tipo_pago_id=4,
            masivo_trans_id__isnull=False,
            cuentatransferencia_id=self._emisor.info_account.get('id')
        ).values_list('masivo_trans_id', flat=True)

        list_sets_massive: List = list(set(massive))
        return list_sets_massive

    @property
    def _list(self) -> Dict[str, Any]:
        l = transmasivaprod.objects.select_related(
            'usuarioRel',
            'statusRel'
        ).filter(
            Q(date_modified__date__gte=self._start_date) &
            Q(date_modified__date__lte=self._end_date)
        ).filter(
            id__in=self._list_id_massive,
            observations__icontains=self._observations,
            statusRel_id=self._status_id
        ).values(
            'id',
            'observations',
            'date_modified',
            'usuarioRel__name',
            'usuarioRel__last_name',
            'date_liberation',
        ).order_by('-date_modified')

        for i in l:
            r = TransMasivaProg.objects.filter(masivaReferida_id=i.get('id')).values('fechaProgramada').first()
            if r:
                i['fechaProgramada'] = r.get('fechaProgramada')

            if not r:
                i['fechaProgramada'] = None
        return l

    def _render_json_transaction(self, **kwargs):
        full_name = f"{kwargs.get('usuarioRel__name')} {remove_asterisk(kwargs.get('usuarioRel__last_name'))}"

        return {
            "id": kwargs.get('id'),
            "observations": kwargs.get('observations'),
            "date_modified": kwargs.get('date_modified'),
            "date_liberation": kwargs.get('date_liberation'),
            "created_to": full_name,
            "fechaProgramada": kwargs.get('fechaProgramada'),
        }


# Endpoint: http://127.0.0.1:8000/transaction/web/v2/LisDisMas/list/
class ListDispersionesMasivasEstado(ListAPIView):
    permission_classes = (BlocklistPermissionV2,)
    permisos = ["Ver dispersiones masivas pendientes", "Ver dispersiones masivas canceladas",
                "Ver dispersiones masivas procesadas"]

    pagination_class = PageNumberPagination

    # size, status_id, observations, date_1, date_2, request.user, razon_social_id
    @method_decorator(cache_page(60 * 0.1))
    def list(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        try:
            admin: persona = request.user
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            log.json_request(request.query_params)
            emisor = GetInfoCompany(admin, **data)
            massive = ClassListDispersionMassive(emisor, **data)
            self.pagination_class.page_size = massive.size
            return self.get_paginated_response(self.paginate_queryset(massive.dispersion_massive))
        except (ObjectDoesNotExist, ValueError) as e:
            err = MyHttpError('Su cuenta no esta asociada a una persona moral', real_error=str(e))
            log.json_response(err.standard_error_responses())
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)


class ClassShowDetailDispersionMassive:
    defaul_size: ClassVar[int] = 5
    detail_dispersion_massive: ClassVar[List[Dict[str, Any]]]

    def __init__(self, emisor: GetInfoCompany, **kwargs):
        self._massive_id = kwargs.get('masivo_trans_id', None)
        self._status_id = kwargs.get('status_id', None)
        self._nombre_beneficiario = kwargs.get('nombre_beneficiario', '')
        self._start_date = kwargs.get('date_1', dt.date.today() - dt.timedelta(days=91))
        self._end_date = kwargs.get('date_2', dt.date.today())
        self.size = kwargs.get('size', self.defaul_size)
        self._emisor = emisor
        self.detail_dispersion_massive = self._list()

    def _raise_params(self) -> NoReturn:
        if not self._status_id and not self._massive_id:
            msg = 'Operación prohibida, debes de enviar el estado de la dispersión y el id de la masiva'
            raise ParamsNotProvided(msg)

        if not self._status_id:
            msg = 'Operación prohibida, debes de enviar el estado de la dispersión'
            raise ParamsNotProvided(msg)

        if not self._massive_id:
            msg = 'Operación prohibida, debes de enviar el estado el id de la masiva'
            raise ParamsNotProvided(msg)

    def _list(self):
        return transferencia.objects.select_related(
            'masivo_trans',
            'status_trans',
            'tipo_pago'
        ).filter(
            Q(date_modify__date__gte=self._start_date) &
            Q(date_modify__date__lte=self._end_date)
        ).filter(
            masivo_trans_id=self._massive_id,
            masivo_trans__statusRel_id=self._status_id,
            cuentatransferencia_id=self._emisor.info_account.get('id'),
            nombre_beneficiario__icontains=self._nombre_beneficiario,
            tipo_pago_id=4
        ).values(
            'id',
            'nombre_beneficiario',
            'monto',
            'fecha_creacion'
        ).order_by('-date_modify')


# Endpoint: http://127.0.0.1:8000/transaction/web/v2/DetDisMas/list/
class DetailDispersionesMassive(ListAPIView):
    def list(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        try:
            admin: persona = request.user
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            log.json_request(request.query_params)
            emisor = GetInfoCompany(admin, **data)
            dispersion_list = ClassShowDetailDispersionMassive(emisor, **data)
            self.pagination_class.page_size = dispersion_list.size
            return self.get_paginated_response(self.paginate_queryset(dispersion_list.detail_dispersion_massive))
        except (ObjectDoesNotExist, ValueError) as e:
            err = MyHttpError('Su cuenta no esta asociada a una persona moral', real_error=str(e))
            log.json_response(err.standard_error_responses())
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)


# __POSIBLE__OBSOLETO
# class DetailDispersionesMassive(ListAPIView):
#     """
#     Autor: Christian Gil
#     Descripción: Regresa el detallado de los listados de dispersión masiva, sin filtrado
#
#     """
#     pagination_class = PageNumberPagination
#     permission_classes = ()
#
#     def get_queryset(self, *args, **kwargs):
#         return filter_data_or_return_none(*args, **kwargs)
#
#     def data_filter(self, beneficiario: str, date_1: datetime, date_2: datetime) -> Dict:
#         return {
#             "beneficiario": '' if beneficiario == 'null' else beneficiario,
#             "date_1": datetime.datetime(2000, 12, 31, 00, 00, 00) if date_1 == 'null' else date_1,
#             "date_2": datetime.datetime.now() if date_2 == 'null' else date_2,
#         }
#
#     def filter_queryset(self, **kwargs) -> List:
#         return transferencia.objects.all().values(
#             'id',
#             'nombre_beneficiario',
#             'monto',
#             'fecha_creacion',
#         ).filter(**kwargs).order_by('fecha_creacion')
#
#     def list(self, request, *args, **kwargs):
#         self.pagination_class.page_size = request.query_params['size']
#         masivo_trans_id: int = self.request.query_params['masivo_trans_id']
#         status_trans_id: int = self.request.query_params['status_id']
#         nombre_beneficiario: str = self.request.query_params['nombre_beneficiario']
#         date_1 = self.request.query_params['date_1']
#         date_2 = self.request.query_params['date_2']
#
#         cuenta_eje: int = get_id_cuenta_eje(request.user.get_only_id())
#         cuenta_id: int = cuenta.objects.get(persona_cuenta_id=cuenta_eje).get_only_id()
#
#         data_filter: Dict = self.data_filter(nombre_beneficiario, date_1, date_2)
#         querysets = self.filter_queryset(
#             masivo_trans_id=masivo_trans_id,
#             status_trans_id=status_trans_id,
#             cuentatransferencia_id=cuenta_id,
#             nombre_beneficiario__icontains=data_filter['beneficiario'],
#             fecha_creacion__range=(data_filter['date_1'], data_filter['date_2']),
#         )
#
#         page = self.paginate_queryset(querysets)
#         return self.get_paginated_response(page)


class ShowDetailsDispersionMasiva(RetrieveAPIView):
    """
    Autor: Christian Gil
    Descripción: Regresa el ver mas detalles del detallado de una dispersión masiva

    """
    serializer_class = ShowDetailSerializer

    def retrieve(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        data = transferencia.objects.get(id=request.query_params['id_dispersion'])
        log.json_request(request.query_params)
        serializer = self.serializer_class(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ComprobanteDisInd(GenericViewSet):
    serializer_class = SerializerDocIndIn

    def create(self, request):
        log = RegisterLog(request.user, request)
        context = {"admin_id": request.user.get_only_id()}
        serializer = self.serializer_class(data=request.data, context=context)
        log.json_request(request.query_params)
        if serializer.is_valid(raise_exception=True):
            documento = serializer.create(serializer.validated_data)
            log.json_response(documento.data)
            return Response(documento.data, status=status.HTTP_200_OK)


class ShowDetailDispersion(RetrieveAPIView):
    def retrieve(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        try:
            transfer_id: int = request.query_params['id']
            log.json_request(request.query_params)
            dispersion_data = transferencia.filter_transaction.detail_dispersion(transfer_id)
            persona = cuenta.objects.get(cuenta=dispersion_data['CuentaBeneficiario'])
            dispersion_data['PropietarioTarjeta'] = persona.persona_cuenta.get_full_name()
            return Response(dispersion_data, status=status.HTTP_200_OK)

        except MultiValueDictKeyError as e:
            err = MyHttpError("Asegurese de haber ingresado todos los parametros", real_error=str(e))
            log.json_response(err.multi_value_dict_key_error())
            return Response(err.multi_value_dict_key_error(), status=status.HTTP_400_BAD_REQUEST)

        except ObjectDoesNotExist as e:
            err = MyHttpError("Dispersión no encontrada", str(e), 404)
            log.json_response(err.object_does_not_exist())
            return Response(err.object_does_not_exist(), status=status.HTTP_404_NOT_FOUND)


class ClassCardSearch:
    cards: ClassVar[List[Dict[str, Any]]]

    def __init__(self, **kwargs):
        self._razon_social_id = kwargs.get('razon_social_id', None)
        self.cards = [self.render(**card) for card in self._list]

    def raise_query_params(self):
        if not self._razon_social_id:
            raise ValueError('Error al buscar la tarjeta')

    @property
    def _list(self) -> List[Dict[str, Any]]:
        return tarjeta.objects.card_search(self._razon_social_id)

    def render(self, **kwargs) -> Dict[str, Any]:
        name = kwargs.get('cuenta__persona_cuenta__name')
        last_name = remove_asterisk(kwargs.get('cuenta__persona_cuenta__last_name'))

        return {
            "nombre": f"{name} {last_name}",
            "email": kwargs.get('cuenta__persona_cuenta__email'),
            "cuenta": kwargs.get('cuenta__cuenta'),
            "tarjeta": kwargs.get('tarjeta')
        }


# Modificado (ChrGil 2021-11-17) Buscar tarjeta en dispersiones individuales
# Endpoint: http://127.0.0.1:8000/transaction/web/v3/CarSeaDis/list/?razon_social_id=1077
class CardSearchDispersion(ListAPIView):
    permission_classes = ()

    def list(self, request, *args, **kwargs):
        log = RegisterLog(request.user, request)
        try:
            data = {key: value for key, value in self.request.query_params.items() if value != 'null'}
            log.json_request(request.query_params)
            lista = ClassCardSearch(**data)
            return Response(lista.cards, status=status.HTTP_200_OK)

        except (MultiValueDictKeyError, ValueError) as e:
            err = MyHttpError("Asegurese de haber ingresado todos los parametros", real_error=str(e))
            log.json_response(err.multi_value_dict_key_error())
            return Response(err.multi_value_dict_key_error(), status=status.HTTP_400_BAD_REQUEST)

        except ObjectDoesNotExist as e:
            err = MyHttpError("Recurso no encontrado", str(e), code=404)
            log.json_response(err.object_does_not_exist())
            return Response(err.object_does_not_exist(), status=status.HTTP_404_NOT_FOUND)


# (Jose Macias 2022-01-03) Crear y descargar Excel de Resporte de cuenta dispersion
class ExcelDispersion(GenericViewSet):
    permission_classes = ()

    def list(self, request):
        log = RegisterLog(request.user, request)
        # Pedif fechas, numnero de cuenta ingreso, egreso o todos
        fecha_inicio = self.request.query_params['FechaInicio']
        fecha_final = self.request.query_params['FechaFinal']
        numero_cuenta = self.request.query_params['NumeroCuenta']
        username_excel = self.request.query_params['Username']
        tipo = self.request.query_params['Tipo']
        log.json_request(request.query_params)
        # tipo
        # - Todos --------> egresos e ingresos
        # - Ingresos 
        # - Egresos 
        # Si las fechas estan Null creamos unas para el filtro
        if fecha_inicio == 'Null' and fecha_final == 'Null':
            fecha_inicio = '1900-01-01'
            fecha_final = str(datetime.date.today())
        # Ordenar fecha como lo pide el Excel
        fecha_inicio_periodo = fecha_inicio
        fecha_final_periodo = fecha_final
        # Ordenar fecha como lo pide la DB
        fecha_inicio = str(fecha_inicio) + ' 00:00:00'
        fecha_final = str(fecha_final) + " 23:59:59"
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, "%Y-%m-%d %H:%M:%S")
        fecha_final = datetime.datetime.strptime(fecha_final, "%Y-%m-%d %H:%M:%S")
        # hacer un query en trasacciones de egresos e ingresos y ordenar por id
        if tipo == 'Todos':
            registros = transferencia.objects.filter(
                Q(cuenta_emisor=numero_cuenta, fecha_creacion__gte=fecha_inicio, fecha_creacion__lte=fecha_final,
                  tipo_pago_id=4) | Q(cta_beneficiario=numero_cuenta, fecha_creacion__gte=fecha_inicio,
                                      fecha_creacion__lte=fecha_final, tipo_pago_id=4)).order_by('id')
        if tipo == 'Ingresos':
            registros = transferencia.objects.filter(cta_beneficiario=numero_cuenta, fecha_creacion__gte=fecha_inicio,
                                                     fecha_creacion__lte=fecha_final, tipo_pago_id=4).order_by('id')
        if tipo == 'Egresos':
            registros = transferencia.objects.filter(cuenta_emisor=numero_cuenta, fecha_creacion__gte=fecha_inicio,
                                                     fecha_creacion__lte=fecha_final, tipo_pago_id=4).order_by('id')

        # hacer un if de todos los filtros

        # Texto final
        texto_final = 'Los recursos de los Usuarios en las operaciones realizadas con Polipay  no se encuentran garantizados por ninguna autoridad. Los fondos de pago electrónico no generan rendimientos o beneficios monetarios por los saldos acumulados en los mismos. Polipay  recibe consultas, reclamaciones o aclaraciones, en su Unidad Especializada de Atención a Usuarios, por correo electrónico a contacto@polipay.com . En el caso de no obtener una respuesta satisfactoria, podrá acudir a la Comisión Nacional para la Protección y Defensa de los Usuarios de Servicios Financieros a través de su página web: https//gob.mx/condusef o al número telefónico 5553400999. '
        # Abri el excel
        excel_estado_cuenta = load_workbook(filename="TEMPLATES/web/EstadoCuenta/Excel/Estado-Cuenta.xlsx")
        sheet = excel_estado_cuenta.active
        # Obtener datos del usuario
        datos_cuenta = cuenta.objects.get(cuenta=numero_cuenta)
        datos_user = persona.objects.get(id=datos_cuenta.persona_cuenta_id)
        # Periodo
        fecha, tiempo = str(fecha_inicio).split(" ")
        if fecha == "1900-01-01":
            fecha_inicio_periodo = registros.first()
            fecha_inicio_periodo, tiempo = str(fecha_inicio_periodo.fecha_creacion).split(" ")
            fecha_final_periodo = registros.last()
            fecha_final_periodo, tiempo = str(fecha_final_periodo.fecha_creacion).split(" ")
            periodo = 'Del ' + str(fecha_inicio_periodo) + ' al ' + str(fecha_final_periodo)
        else:
            periodo = 'Del ' + str(fecha_inicio_periodo) + ' al ' + str(fecha_final_periodo)
        sheet['G18'] = periodo
        # Colocar datos del usuario en el excel
        sheet['B11'] = datos_user.name
        sheet['B11'].font = Font(u'Arial', bold=True, size=11)
        sheet['C13'] = datos_cuenta.cuenta
        sheet['C14'] = datos_user.id
        sheet['F11'] = datos_user.name
        try:
            direccion = domicilio.objects.get(domicilioPersona_id=datos_user.id)
            domicilio_user = direccion.calle + ' ' + direccion.no_exterior + ', COL ' + direccion.colonia \
                             + ', DELEG ' + direccion.alcaldia_mpio + ', CP ' + direccion.codigopostal + ', ' + direccion.estado
            sheet['F12'] = domicilio_user
        except:
            pass
        numero = 31

        numero_abonos_depositos = 0
        total_abonos_depositos = 0
        numero_retiros_cargos = 0
        total_retiros_cargos = 0
        saldo_anterior = registros.first()
        saldo_final = registros.last()
        # Colocamos saldo anterior y saldo final
        sheet['D20'] = saldo_anterior.saldo_remanente
        sheet['D23'] = saldo_final.saldo_remanente
        # B ---> Fecha de operacion
        # C ---> Fecha de liquidacion
        # D ---> concepto
        # E ---> Clave de rastreo
        # F ---> Cargos
        # G ---> Abonos
        # H ---> Saldo
        # hacer un for de la trasferencia
        for registro in registros:
            # Acomodar la fecha para el excel
            fecha_creacion, tiempo = str(registro.fecha_creacion).split(' ')
            # Acomodar los datos en las celdas
            sheet["B" + str(numero)] = fecha_creacion
            sheet['B' + str(numero)].alignment = Alignment(horizontal="left")
            sheet["C" + str(numero)] = registro.date_modify
            sheet['C' + str(numero)].alignment = Alignment(horizontal="left")
            sheet["D" + str(numero)] = registro.concepto_pago
            sheet['D' + str(numero)].alignment = Alignment(horizontal="left")
            sheet["E" + str(numero)] = registro.clave_rastreo
            sheet['E' + str(numero)].alignment = Alignment(horizontal="left")
            # Ordenar el monto con centavos
            locale.setlocale(locale.LC_ALL, 'en_US')
            # if '.' in str(registro.monto):
            #     print(registro.id)
            #     print('tiene punto decimal')
            saldo, centavos = str(registro.monto).split(".")
            # else:
            #     print('no tiene punto decimal')
            #     saldo = registro.monto
            #     centavos = '00'
            # Ver si es egreso o ingreso y acomodarlo 
            if numero_cuenta == registro.cuenta_emisor:
                sheet["F" + str(numero)] = "$" + str(f'{int(saldo):n}') + '.' + str(centavos[:2])
                sheet['F' + str(numero)].alignment = Alignment(horizontal="left")
                numero_retiros_cargos = numero_retiros_cargos + 1
                total_retiros_cargos = float(total_retiros_cargos) + float(registro.monto)
            else:
                sheet["G" + str(numero)] = "$" + str(f'{int(saldo):n}') + '.' + str(centavos[:2])
                sheet['G' + str(numero)].alignment = Alignment(horizontal="left")
                numero_abonos_depositos = numero_abonos_depositos + 1
                total_abonos_depositos = float(total_abonos_depositos) + float(registro.monto)
            # Ordenar saldo
            # if registro.saldo_remanente is '.':
            #     saldo, centavos = str(registro.saldo_remanente).split(".")
            # else:
            #     saldo = registro.saldo_remanente
            #     centavos = '00'
            try:
                saldo_remanente = float(registro.saldo_remanente)
                saldo, centavos = str(saldo_remanente).split(".")
            except:
                saldo, centavos = '0', '000'
            try:
                sheet["H" + str(numero)] = '$' + str(f'{int(saldo):n}') + '.' + str(centavos[:2])
                sheet['H' + str(numero)].alignment = Alignment(horizontal="left")
            except:
                sheet["H" + str(numero)] = '$0.00'
                sheet['H' + str(numero)].alignment = Alignment(horizontal="left")
            numero = int(numero) + 1
        numero_inicio = numero + 2
        numero_final = numero + 4
        sheet.merge_cells('B' + str(numero_inicio) + ':H' + str(numero_final))
        cell = sheet.cell(row=numero_inicio, column=2)
        cell.value = texto_final
        cell.alignment = cell.alignment.copy(wrapText=True)  # ----> Sirve para ajustar el texto en las celdas
        # Colocamos cantidad de abonos/depositos y retiros/cargos
        sheet['C21'] = numero_abonos_depositos
        sheet['C22'] = numero_retiros_cargos
        # Colocamos total de depositos y cargos
        sheet['D21'] = total_abonos_depositos
        sheet['D22'] = total_retiros_cargos
        fecha_actual = str(datetime.date.today()).replace('-', '')
        excel_estado_cuenta.save(
            filename='TMP/web/Estado_Cuentas/Excel/Estado-cuenta-Dispersion-' + username_excel + '_' + fecha_actual + ".xlsx")
        # Descargar Excel
        filename = 'TMP/web/Estado_Cuentas/Excel/Estado-cuenta-Dispersion-' + username_excel + '_' + fecha_actual + ".xlsx"
        filepath = filename
        path = open(filepath, 'r')
        mime_type, _ = mimetypes.guess_type(filepath)
        response = FileResponse(open(filename, 'rb'))
        response['Content-Disposition'] = "attachment; filename=%s" % filename

        return Response({'status': 'pruebas'}, status=status.HTTP_200_OK)


class ComponentInfoCostCenter:
    def __init__(self, cost_center_id: int):
        self._cost_center_id = cost_center_id
        emisor = self._info_emisor
        cuenta_eje_name = self._cuenta_eje_name

        if emisor:
            self.emisor = emisor

        if cuenta_eje_name:
            self.cuenta_eje_name = cuenta_eje_name

        if not emisor:
            raise ValueError("Centro de costos no encontrado o no existe")

    @property
    def _info_emisor(self) -> Dict[str, Any]:
        return cuenta.objects.get_info_with_user_id(owner=self._cost_center_id)

    @property
    def _cuenta_eje_name(self) -> Dict[str, Any]:
        return grupoPersona.objects.get_name_cuenta_eje(person_id=self._cost_center_id)


def _add_saldo_remanente(transaction: transferencia, current_amount: float, amount_transaction: float):
    saldo_remanente = float(current_amount)
    saldo_remanente += float(amount_transaction)
    transaction.saldo_remanente_beneficiario = saldo_remanente
    transaction.save()


class DevolverComissionToPolipayComission:
    """ Pago por SPEI de una comisión devuelta """

    _registra_orden: ClassVar[RegistraOrdenDispersionMasivaIndividual] = RegistraOrdenDispersionMasivaIndividual
    _rs_polipay_comission: ClassVar[int] = COST_CENTER_POLIPAY_COMISSION

    def __init__(self, emisor: ComponentInfoCostCenter, total_amount: float, log: RegisterLog):
        self._emisor = emisor
        self._total_comission = total_amount
        self._comission_info = self._get_info_polipay_comission
        instance = self.create(**self._data)
        self._registra_orden(instance, log)
        cuenta.objects.withdraw_amount(self._comission_info.get("persona_cuenta_id"), amount=total_amount)
        _add_saldo_remanente(instance, total_amount, emisor.emisor.get("monto"))

    @property
    def _get_info_polipay_comission(self) -> Dict[str, Any]:
        return cuenta.objects.get_info_polipay_comission(self._rs_polipay_comission)

    @property
    def _data(self) -> Dict[str, Any]:
        return {
            "empresa": self._comission_info.get("persona_cuenta__name_stp"),
            "rfc_curp_emisor": self._comission_info.get("persona_cuenta__rfc"),
            "nombre_emisor": self._comission_info.get("persona_cuenta__name"),
            "cuenta_emisor": self._comission_info.get("cuentaclave"),
            "cuentatransferencia_id": self._comission_info.get("id"),
            "monto": self._total_comission,
            "nombre_beneficiario": self._emisor.emisor.get("persona_cuenta__name"),
            "cta_beneficiario": self._emisor.emisor.get("cuentaclave"),
            "rfc_curp_beneficiario": self._emisor.emisor.get("persona_cuenta__rfc"),
            "concepto_pago": "Comisión devuelta",
            "referencia_numerica": "1234567"
        }

    @staticmethod
    def create(**kwargs) -> transferencia:
        return transferencia.objects.create_transaction_polipay_to_polipay_v2(**kwargs, tipo_pago_id=11)


class DevuelveDineroSaldosWallet:
    """ El ceco Saldos Wallet regresa por SPEI el monto de la dispersión si es cancelada """

    _registra_orden: ClassVar[RegistraOrdenDispersionMasivaIndividual] = RegistraOrdenDispersionMasivaIndividual
    _rs_saldos_wallet: ClassVar[int] = COST_CENTER_INNTEC

    def __init__(self, emisor: ComponentInfoCostCenter, total_amount: float, log: RegisterLog):
        self._emisor = emisor
        self._total_amount = total_amount
        self._saldos_wallet_info = self._get_info_saldos_wallet
        instance = self.create(**self._data)
        self._registra_orden(instance, log)
        cuenta.objects.withdraw_amount(self._saldos_wallet_info.get("persona_cuenta_id"), amount=total_amount)
        _add_saldo_remanente(instance, total_amount, emisor.emisor.get("monto"))

    @property
    def _get_info_saldos_wallet(self) -> Dict[str, Any]:
        return cuenta.objects.get_info_with_user_id(self._rs_saldos_wallet)

    @property
    def _data(self) -> Dict[str, Any]:
        return {
            "empresa": self._saldos_wallet_info.get("persona_cuenta__name_stp"),
            "rfc_curp_emisor": self._saldos_wallet_info.get("persona_cuenta__rfc"),
            "nombre_emisor": self._saldos_wallet_info.get("persona_cuenta__name"),
            "cuenta_emisor": self._saldos_wallet_info.get("cuentaclave"),
            "cuentatransferencia_id": self._saldos_wallet_info.get("id"),
            "monto": self._total_amount,
            "nombre_beneficiario": self._emisor.emisor.get("persona_cuenta__name"),
            "cta_beneficiario": self._emisor.emisor.get("cuentaclave"),
            "rfc_curp_beneficiario": self._emisor.emisor.get("persona_cuenta__rfc"),
            "concepto_pago": "Monto Devuelto",
            "referencia_numerica": "1234567",
        }

    @staticmethod
    def create(**kwargs) -> transferencia:
        return transferencia.objects.create_transaction_polipay_to_polipay_v2(**kwargs, tipo_pago_id=1)


# (ManuelCalixtro 06/06/2022) Endpoint para cancelar una dispersion programada individual
class CancelDispersionSheduled(GenericViewSet):
    _info_cost_center: ClassVar[ComponentInfoCostCenter] = ComponentInfoCostCenter
    _devuelve_commision: ClassVar[DevolverComissionToPolipayComission] = DevolverComissionToPolipayComission
    _devolver_monto_saldo: ClassVar[DevuelveDineroSaldosWallet] = DevuelveDineroSaldosWallet

    def create(self):
        pass

    def put(self, request):
        log = RegisterLog(request.user, request)

        try:
            with atomic():
                log.json_request(self.request.query_params)
                dispersion_id = self.request.query_params['dispersion_id']
                clave_rastreo = self.request.query_params['clave_rastreo']
                cost_center_id = self.request.query_params['cost_center_id']
                emisor = self._info_cost_center(cost_center_id)

                transferencia.objects.filter(
                    clave_rastreo=clave_rastreo,
                    programada=True,
                    tipo_pago_id=4
                ).update(
                    status_trans_id=5,
                    date_modify=datetime.datetime.now()
                )

                transferenciaProg.objects.filter(transferReferida_id=dispersion_id).delete()

                instance_trasnferencia = transferencia.objects.get(id=dispersion_id, tipo_pago_id=4, status_trans_id=5)
                instance_cuenta = cuenta.objects.get(Q(cuenta=instance_trasnferencia.cuenta_emisor) | Q(
                    cuentaclave=instance_trasnferencia.cuenta_emisor))
                # instance_cuenta.monto += instance_trasnferencia.monto
                # instance_cuenta.save()

                instance_comission = Commission_detail.objects.get(transaction_rel_id=dispersion_id)
                instance_comission.status_id = 5
                instance_comission.save()

                # return_comision = transferencia.objects.create_transaction_to_return_comission(
                #     instance_comission.mount,
                #     instance_cuenta.persona_cuenta.name,
                #     instance_cuenta.cuenta,
                #     instance_cuenta.persona_cuenta.rfc,
                # )

                # instance_cuenta.monto += float(instance_comission.mount)
                # instance_cuenta.save()

                # return_comision.saldo_remanente_beneficiario += float(instance_cuenta.monto)
                # return_comision.save()

                # instance_polipay_comission = cuenta.objects.get(persona_cuenta_id=COST_CENTER_POLIPAY_COMISSION)
                # instance_polipay_comission.monto -= float(instance_comission.mount)
                # instance_polipay_comission.save()

                # Regresa monto por SPEI de la comisión al centro de costos que cancela la operación
                self._devuelve_commision(emisor, instance_comission.mount, log)
                self._devolver_monto_saldo(emisor, instance_trasnferencia.monto, log)

        except (ObjectDoesNotExist, IntegrityError, ValueError, TypeError) as e:
            err = MyHttpError(message="Ocurrió un error al momento de cancelar la dispersion", real_error=str(e))
            log.json_response(err.standard_error_responses())
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)
        except StpmexException as e:
            message = "Ocurrio un error al cancelar la operación"
            err = MyHttpError(message=message, real_error=e.desc)
            log.json_response(err.standard_error_responses())
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)
        else:
            succ = MyHtppSuccess(message='Tu operación se realizo de manera satisfactoria', code='200')
            log.json_response(succ.standard_success_responses())
            return Response(succ.standard_success_responses(), status=status.HTTP_201_CREATED)
