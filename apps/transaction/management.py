import io
import uuid
import locale
import random
import base64
import datetime
from typing import Dict, Any
import pandas as pde

from openpyxl import load_workbook
from PyPDF2 import PdfFileWriter, PdfFileReader

from django.core.files import File
from django.db import connection

from rest_framework.serializers import ValidationError
from rest_framework.exceptions import ParseError
from rest_framework.response import Response
from rest_framework import status

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import HexColor
from textwrap import wrap

from apps.transaction.messages import createMessageTransactionMassive
from apps.transaction.api.movil.serializers.createTransaction import *
from apps.users.models import documentos, persona
from apps.transaction.models import transferencia
from MANAGEMENT.notifications.movil.notifyAppUser import notifyAppUserFromWeb

leyend = 'Los recursos de los Usuarios en las operaciones realizadas con Polipay  no se encuentran garantizados por ninguna autoridad. ' \
         'Los fondos de pago electrónico no generan rendimientos o beneficios monetarios por los saldos acumulados en los mismos. ' \
         'Polipay  recibe consultas, reclamaciones o aclaraciones, en su Unidad Especializada de Atención a Usuarios, por correo electrónico a contacto@polipay.com . En el caso de no obtener una respuesta satisfactoria, podrá acudir a la Comisión Nacional para la Protección y ' \
         'Defensa de los Usuarios de Servicios Financieros a través de su página web: https//gob.mx/condusef o al número telefónico 5553400999. '


def createExcelData(data):
    decrypted = base64.b64decode(data)
    with open("Files/file.xlsx", "wb") as f:
        f.write(decrypted)
    return True


def checkNan(data):
    if pde.isna(data):
        return None
    else:
        return data


#__POSIBLE__OBSOLETO
def ceateTransactionIndividualMasive(datos, instanceM, IdPersona, IdStatus, Id_account):
    dateNow = datetime.datetime.now()
    try:
        instance = transferencia()
        instance.cta_beneficiario = datos[0]
        instance.bancos = datos[1]
        instance.clave_rastreo = datos[2]
        instance.nombre_beneficiario = datos[3]
        instance.rfc_curp_beneficiario = datos[4]
        instance.tipo_pago_id = datos[5]
        instance.tipo_cuenta = datos[6]
        instance.monto = datos[7]
        instance.concepto_pago = datos[8]
        instance.referencia_numerica = datos[9]
        instance.institucion_operante = datos[10]
        instance.masivo_trans_id = instanceM.id
        instance.cuentatransferencia_id = Id_account.id
        instance.cuenta_emisor = Id_account.cuenta
        instance.date_modify = dateNow
        createMessageTransactionMassive(instanceM, IdPersona, IdStatus)
        instance.save()
    except Exception as inst:
        instanceM.delete()
        raise ParseError({'status': 'Hubo un error', "error": inst})
    return True


def getReferenceNumber():
    cuenta = random.random()
    cuenta = cuenta * 100000000
    cuenta = hex(int(cuenta))
    return cuenta.upper()


def getClaveRastrero(data):  # dataes el nombre del receptor
    cuenta = int(random.random() * 10000000)
    cuenta = hash(cuenta)
    if cuenta < 0:
        cuenta = cuenta * -1
    cuenta = hex(cuenta)
    key = hash(data)
    if key < 0:
        key = key * -1
    key = hex(key)
    clave = "PO" + key + str(cuenta)
    clave = clave.upper()
    return clave[0:30]


def make_transanction(data):
    with connection.cursor() as cursor:
        result = cursor.callproc('Transferencias', data)
    return result


def getKward(request):
    diccionario = {}

    if request['date_start'] != 'null':
        date_s = datetime.datetime.strptime(request['date_start'], '%d/%m/%y %H:%M:%S')
        diccionario['fecha_creacion__gte'] = date_s.date()
    if request['date_start'] == 'null':
        datenow = datetime.datetime.now()
        date_s = datenow - datetime.timedelta(days=30)
        diccionario['fecha_creacion__gte'] = date_s.date()
    if request['date_end'] == 'null':
        date_e = datetime.datetime.now()
        diccionario['fecha_creacion__lte'] = date_e.date()
    if request['date_end'] != 'null':
        date_e = datetime.datetime.strptime(request['date_end'], '%d/%m/%y %H:%M:%S')
        diccionario['fecha_creacion__lte'] = date_e.date()

    return diccionario


def get_filter_data(instance, *args, **kwargs):
    return instance.objects.filter(*args, **kwargs)


#__POSIBLE__OBSOLETO
def to_string_date(date, format=None):
    if format is None:
        return datetime.datetime.strftime(date, '%d/%m/%y %H:%M')
    return datetime.datetime.strftime(date, format)


def to_base64_file(file):
    with open(file, 'rb') as binary_file:
        binary_file_data = binary_file.read()
        base64_encoded_data = base64.b64encode(binary_file_data)
        return base64_encoded_data.decode('utf-8')


#__POSIBLE__OBSOLETO
def save_excel_report(grupo_persona_id, file):
    """ Recibe un id y un objeto y regresa un base64_file """

    fileDate = to_string_date(datetime.datetime.now(), '%d%m%Y')
    # empresa = grupoPersona.objects.filter(person_id=grupo_persona_id).first()
    file_name = f'FileSystem/{fileDate}.xlsx'
    file.save(file_name)
    return to_base64_file(file_name)


#__POSIBLE__OBSOLETO
def write_excel_report(instance):
    global moral
    file = load_workbook('FileSystem/ExcelFile.xlsx')
    ws = file.get_sheet_by_name("hoja1")
    row = 1

    for i in instance:
        row += 1
        ws.cell(column=1, row=row, value=i.id)
        ws.cell(column=2, row=row, value=i.empresa)
        ws.cell(column=3, row=row, value=i.cuentatransferencia_id)
        ws.cell(column=4, row=row, value=i.banco_emisor)
        ws.cell(column=5, row=row, value=i.clave_rastreo)
        ws.cell(column=6, row=row, value=i.nombre_beneficiario)
        ws.cell(column=7, row=row, value=i.rfc_curp_beneficiario)
        ws.cell(column=8, row=row, value=i.email)
        ws.cell(column=9, row=row, value=i.tipo_pago.nombre_tipo)
        ws.cell(column=10, row=row, value=i.tipo_cuenta)
        ws.cell(column=11, row=row, value=i.monto)
        ws.cell(column=12, row=row, value=i.concepto_pago)
        ws.cell(column=13, row=row, value=i.referencia_numerica)
        ws.cell(column=14, row=row, value=i.institucion_operante)
        ws.cell(column=15, row=row, value=to_string_date(i.fecha_creacion))
        ws.cell(column=16, row=row, value=to_string_date(i.date_modify))
        moral = i.cuentatransferencia_id

    ws.cell(column=1, row=row + 5, value=leyend)
    return save_excel_report(moral, file)


#__POSIBLE__OBSOLETO
def transaction_status_change(query_transaction, instance_cuenta, status, instance_persona, message):
    query_transaction.status_trans_id = status
    query_transaction.date_modify = datetime.datetime.now()
    query_transaction.save()

    if status == 1:
        return True
    if status == 2 or 5:
        instance_cuenta.monto += query_transaction.monto
    if status == 3:
        instance_cuenta.monto -= query_transaction.monto
    instance_cuenta.save()
    # createMessageTransactionSend(instance_persona, query_transaction, message)
    return Response({"status": f'¡Transferencia cambiada a estado {message} de forma exitosa!'})


def CrearComprobanteDispersionPDF(instance_disper, beneficiario, userId):
    date, time = str(instance_disper.fecha_creacion).split(" ")
    dateCrea, time1 = str(instance_disper.fecha_creacion).split(".")
    # date = datetime.datetime.strptime(str(instance_disper.fecha_creacion), '%Y-%m-%d')
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    ### Escribimos en el PDF
    can.drawString(115, 543, str(instance_disper.id))  # No. operacion
    can.drawString(95, 515.5, instance_disper.nombre_emisor)  # Ordenante
    can.drawString(100, 488, beneficiario)  # Beneficiario Nombre
    can.drawString(150, 461.5, instance_disper.cta_beneficiario)  # Cuenta Dispersion es beneficiario
    can.drawString(80, 433.5, f'$' + str(instance_disper.monto))  # importe
    can.drawString(90, 405, instance_disper.concepto_pago)  # concepto
    can.drawString(145, 378, str(dateCrea))  # fecha Operacion
    can.drawString(460, 542.5, str(date))  # Fecha
    can.save()
    ### Creamos el PDF
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    existing_pdf = PdfFileReader(open("TEMPLATES/web/ComprobanteDispersion.pdf", "rb"))
    output = PdfFileWriter()
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    outputStream = open("TMP/web/ComprobanteDispersionCompleto.pdf", "wb")
    output.write(outputStream)
    outputStream.close()
    documentoId = SubirDocumentoDispersionIndividual(userId)
    return documentoId


def SubirDocumentoDispersionIndividual(userId):
    with open('TMP/web/ComprobanteDispersionCompleto.pdf', 'rb') as document:
        instance = documentos.objects.filter(person_id=userId, tdocumento_id=18, historial=False)
        for inst in instance:
            inst.historial = True
            inst.save()
        instance_document = documentos.objects.create(person_id=userId,
                                                      comentario="Documento de dispersion individual",
                                                      documento="",
                                                      tdocumento_id=18, historial=False)

        instance_document.documento = File(document)
        instance_document.save()
    return instance_document.id


def CrearComprobanteTransactionPDF(instance_disper, beneficiario, userId):
    # date, time = str(instance_disper.fecha_creacion).split(" ")
    # date, time = str(instance_disper.fecha_creacion).split(" ")
    date_created = datetime.datetime.strftime(instance_disper.date_modify, '%Y-%m-%d %H:%M:%S')
    date = datetime.datetime.strftime(instance_disper.fecha_creacion, '%Y-%m-%d')
    packet = io.BytesIO()

    can = canvas.Canvas(packet, pagesize=letter)
    ### Escribimos en el PDF
    can.drawString(165, 542, str(instance_disper.id))  # No. operacion
    can.drawString(150, 515.5, instance_disper.nombre_emisor)  # Ordenante
    can.drawString(195, 488, instance_disper.cuenta_emisor)  # Cuenta Dispersion ordenante
    can.drawString(155, 461, beneficiario)  # Beneficiario Nombre
    can.drawString(165, 432, instance_disper.cta_beneficiario)  # Cuenta Dispersion es beneficiario
    can.drawString(135, 405, instance_disper.receiving_bank.institucion)  # Banco beneficiario
    can.drawString(130, 378, instance_disper.status_trans.nombre)  # Estado transaccion
    # can.drawString(135, 350, f'$' + str(instance_disper.monto))  # importe
    can.drawString(135, 350, f'{"$"}{instance_disper.monto:3,.2f}')   # importe
    can.drawString(185, 323, instance_disper.referencia_numerica)  # referencia numerica
    can.drawString(185, 296, instance_disper.clave_rastreo)  # referencia numerica
    can.drawString(145, 269, instance_disper.concepto_pago) # concepto
    can.drawString(195, 241.5, str(date_created))  # fecha Operacion
    can.drawString(460, 542.5, str(date))  # Fecha
    can.save()
    ### Creamos el PDF
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    existing_pdf = PdfFileReader(open("TEMPLATES/web/Comprobante.pdf", "rb"))
    output = PdfFileWriter()
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    outputStream = open("TMP/web/ComprobanteTransaccion.pdf", "wb")
    output.write(outputStream)
    outputStream.close()
    documentoId = SubirDocumentoComprobanteTransaccion(userId)
    return documentoId


def SubirDocumentoComprobanteTransaccion(userId):
    with open('TMP/web/ComprobanteTransaccion.pdf', 'rb') as document:
        instance = documentos.objects.filter(person_id=userId, tdocumento_id=18, historial=False)
        for inst in instance:
            inst.historial = True
            inst.save()
        instance_document = documentos.objects.create(person_id=userId,
                                                      comentario="Comprobante Transaccion",
                                                      documento="",
                                                      tdocumento_id=23, historial=False)

        instance_document.documento = File(document)
        instance_document.save()
    return instance_document.id

## Calcylar hojas

def CalcularHojasTotales(datos):
    n = 1
    if datos <= 8:
        return 1
    if datos <= 30:
        return 2
    else:
        n = n + 2
        for i in range(44, 10000000000000000000000, 22):
            if datos <= i:
                return n
            else:
                n = n + 1


def PDFMasivas(queryset, username, masivoId):
    errores = []
    try:
        periodo = transmasivaprod.objects.get(id=masivoId)
        query = queryset.last()
        cuentaUser = cuenta.objects.get(cuenta=query.cuenta_emisor)
        usuario = persona.objects.get(id=cuentaUser.persona_cuenta_id)
        domicilioUser = domicilio.objects.get(domicilioPersona_id=usuario.id)
        servicio = tipo_transferencia.objects.get(id=query.tipo_pago_id)
        nRetiros = 0
        retiros = 0.0
        for i in queryset:
            retiros = retiros + i.monto
            nRetiros = nRetiros + 1
        datos = queryset.count()
        numPagTotal = CalcularHojasTotales(datos)
        pdf_file = 'TMP/web/movimientos-DispMassiva' + username + ".pdf"
        carpetaImgDisMas = "TEMPLATES/web/DispMasivasImg/"
        can = canvas.Canvas(pdf_file, pagesize=letter)
        can.drawImage(carpetaImgDisMas + "Dispersion-masiva-y-logo.jpg", 0, 708, width=614, height=84)
        can.drawImage(carpetaImgDisMas + "informacion-periodo.jpg", 362, 610, width=235, height=45)
        can.drawImage(carpetaImgDisMas + "informacion-financiera.jpg", 28, 395, width=290, height=127)
        can.drawImage(carpetaImgDisMas + "detalles-movimientos-realizados.jpg", 5, 304, width=607, height=92)
        can.drawImage(carpetaImgDisMas + "pie-pagina.jpg", 5, 2, width=607, height=50)
        can.setFontSize(8.5)
        can.setFillColor("black")
        can.drawString(30, 655, usuario.name)  # empresa
        can.drawString(30, 642, servicio.nombre_tipo)  # nombre del servicio ejemplo Polipay Dispersa
        can.setFillColor(HexColor('#5c5a59'))
        can.drawString(30, 630, "No. Cuenta: " + str(cuentaUser.cuenta))  # NO. cuenta
        can.drawString(30, 617, "No. Cliente / Prefijo: " + str(usuario.id))  # NO.CLIENTE/b PREFiJO
        can.drawString(30, 605, "RFC: " + usuario.rfc)  # RFC
        can.drawString(30, 585, domicilioUser.calle + " " + domicilioUser.no_exterior)  # Calle y numero
        can.drawString(30, 570, domicilioUser.colonia)  # Colonia
        can.drawString(30, 555, domicilioUser.alcaldia_mpio)  # Delegacion
        can.drawString(30, 540, "CP " + domicilioUser.codigopostal + " " + domicilioUser.estado)  # CP y Estado
        can.drawString(420, 623, periodo.observations)  # Periodo
        can.setFontSize(8.4)
        locale.setlocale(locale.LC_ALL, 'en_US')
        SA = queryset.values().last()
        SF = queryset.values().first()
        SaldoA, CenA = str(SA.get("saldo_remanente")).split(".")
        SaldoF, CenF = str(SF.get("saldo_remanente")).split(".")
        Retiros, CenR = str(retiros).split(".")
        can.drawRightString(305, 475,
                            "$" + str(f'{int(SaldoA):n}' + "." + CenA))  # saldo anterior #primero que se realizo
        can.drawRightString(305, 415, "$" + str(f'{int(SaldoF):n}' + "." + CenF))  # Saldo final #ultimo que se realizo
        can.drawCentredString(180, 435, str(nRetiros))  # Retiros / Cargos #
        can.setFillColor("red")
        can.drawRightString(305, 435, "$" + str(f'{int(Retiros):n}' + "." + CenR))  # Retiros / Cargos
        pagina = 1
        can.setFillColor("gray")
        can.setFontSize(9.1)
        can.drawRightString(600, 20, str(pagina) + "/" + str(numPagTotal))
        can.setLineWidth(0)
        lineY = 276
        lineYD = 295
        lineXFO = 80
        lineXB = 218
        lineXN = 334.5
        lineXD = 453
        lineXM = 570
        for datos in queryset:
            if lineY < 66 and lineYD < 80:
                can.showPage()
                can.setLineWidth(0)
                lineY = 696
                lineYD = 715
                can.drawImage(carpetaImgDisMas + "detalles-movimientos-realizados.jpg", 5, 725, width=607, height=92)
                can.drawImage(carpetaImgDisMas + "pie-pagina.jpg", 5, 2, width=607, height=50)
                can.setFillColor("gray")
                can.setFontSize(9.1)
                pagina = int(pagina) + 1
                can.drawRightString(600, 20, str(pagina) + "/" + str(numPagTotal))
            can.setLineWidth(0)
            can.setFontSize(8.4)
            can.setFillColor(HexColor('#5c5a59'))
            fecha1, segundos = str(datos.fecha_creacion).replace("-", "/").split(".")
            tiempo = "am"
            fecha, hora = fecha1.replace(":", "").split(" ")
            if int(hora) > 120000:
                tiempo = "pm"
            can.drawCentredString(lineXFO, lineYD, fecha1 + " " + tiempo)  # fecha
            beneficiario = datos.nombre_beneficiario  # Beneficiario
            text = "^".join(wrap(beneficiario, 28))
            if "^" in text:
                parrafo1, parrafo2 = text.split("^")
                can.drawCentredString(lineXB, lineYD, parrafo1)
                lineYDN = lineYD - 10
                can.drawCentredString(lineXB, lineYDN, parrafo2)
            else:
                can.drawCentredString(lineXB, lineYD, beneficiario)
            can.drawCentredString(lineXN, lineYD, datos.cta_beneficiario)  # Num cuenta
            descripcion = datos.concepto_pago
            text = "^".join(wrap(descripcion, 28))
            if "^" in text:
                parrafo1, parrafo2 = text.split("^")
                can.drawCentredString(lineXD, lineYD, parrafo1)
                lineYDN = lineYD - 10
                can.drawCentredString(lineXD, lineYDN, parrafo2)
            else:
                can.drawCentredString(lineXD, lineYD, descripcion)
            Monto, CenM = str(datos.monto).split(".")
            can.drawCentredString(lineXM, lineYD, "$" + str(f'{int(Monto):n}' + "." + CenM))  # Monto
            can.setLineWidth(0)
            can.line(28.3, lineY, 600, lineY)
            lineY = lineY - 30
            lineYD = lineYD - 30
        can.save()
        return
    except:
        errores.append(
            {"field": "", "data": "",
             "message": "No se pudo generar el pdf por falta de información de la transferencia masiva"})
        raise ValidationError({"Error": {"code": ["400"]}, "status": ["ERROR"], "detail": [errores]})


# # (ChrGil 2021-10-13) De la fecha actual, se le restan 3 meses y se regresa la fecha resultante
# def get_last_three_months(date):
#     current_year: int = date.year
#     print(current_year)
#     last_three_months: int = date.month - 3
#     print(last_three_months)
#     day: int = date.day
#     return datetime.datetime(current_year, last_three_months, day)


# (ChrGil 2021-10-13) Depenediendo del valor del parametro, se realizara una tarea distintan
def to_dict_query_params(query_params) -> Dict:
    data: Dict[str, Any] = {}
    for key, value in query_params.items():
        if value == 'null':
            if key == 'start_date':
                new_date = datetime.datetime.now() - datetime.timedelta(days=91)
                data[key] = new_date
                continue

            if key == 'end_date':
                data[key] = datetime.datetime.now()
                continue

            data[key] = ''
        if value != 'null':
            data[key] = value

    data.pop('size')
    return data



def preparingNotification(idDeTransferencia=0, cuentaBeneficiario="", idPersona=0, opcion=0):
    # Esta función recupera los datos de la trasnferencia del emisor y beneficiario.
    # Hasta el momento (Jue 02.12.2021) el cambio de estado a LIQUIDADA, DEVUELTA son para las trasnferencias
    # realizadas desde la Wallet (Polipay a Terceros).

    # Para escenario 4, opcion=3
    pk_cuenta = 0

    # Escenario en el que se realicen los preparativos para RECIBIDAS
    if opcion == 2:
        queryExistePersona = persona.objects.filter(id=idPersona).exists()
        if not queryExistePersona:
            return Response({"status": "Usuario no existe o no pertenece\na Polipay, favor de verificar los datos."},
                            status=status.HTTP_400_BAD_REQUEST)

        queryIdCuenta = cuenta.objects.filter(persona_cuenta_id=idPersona).values("id")
        if len(queryIdCuenta) == 0 or queryIdCuenta  == None or queryIdCuenta  == "":
            return Response({"status": "La cuenta no existe o no pertenece\na Polipay, favor de verificar los datos."},
                            status=status.HTTP_400_BAD_REQUEST)

        pk_cuenta   = queryIdCuenta[0]["id"]

        queryIdTransf   = transferencia.objects.filter(cuentatransferencia_id=pk_cuenta, tipo_pago_id=5).values("id").last()
        if len(queryIdTransf) == 0 or queryIdTransf == None or queryIdTransf == "":
            return Response({"status": "Transferencia para el beneficiario no existe o no pertenece\na Polipay, "
                "favor de verificar los datos."}, status=status.HTTP_400_BAD_REQUEST)

        idDeTransferencia   = queryIdTransf["id"]


    # Escenario en el que se realicen los preparativos para RECIBIDAS (DISPERSION, ej: Pago de nomina)
    if opcion == 3:
        queryIdCuenta = cuenta.objects.filter(cuenta=cuentaBeneficiario).values("id")
        if len(queryIdCuenta) == 0 or queryIdCuenta == None or queryIdCuenta == "":
            return Response({"status": "La cuenta del beneficiario no existe o no pertenece\na Polipay, favor de verificar los datos."},
                            status=status.HTTP_400_BAD_REQUEST)

        pk_cuenta = queryIdCuenta[0]["id"]

        queryIdTransf = transferencia.objects.filter(cta_beneficiario=cuentaBeneficiario,tipo_pago_id=4).values("id").last()
        if len(queryIdTransf) == 0 or queryIdTransf == None or queryIdTransf == "":
            return Response({"status": "Transferencia para el beneficiario no existe o no pertenece\na Polipay, "
                                       "favor de verificar los datos."}, status=status.HTTP_400_BAD_REQUEST)

        idDeTransferencia = queryIdTransf["id"]


    queryExisteTransf   = transferencia.objects.filter(id=idDeTransferencia).exists()
    if not queryExisteTransf:
        return Response({"status": "La transferencia no existe, favor de verificar los datos."},
                        status=status.HTTP_400_BAD_REQUEST)

    queryTransferencia  = transferencia.objects.filter(id=idDeTransferencia).values("id", "tipo_pago_id",
        "status_trans_id", "monto", "nombre_emisor", "cuenta_emisor", "nombre_beneficiario", "cta_beneficiario",
        "referencia_numerica", "cuentatransferencia_id", "empresa")
    if len(queryTransferencia) == 0 or queryTransferencia == None or queryTransferencia == "":
        return Response({"status": "Cuenta beneficiario no existe o no pertenece\na Polipay, favor de verificar los datos."},
                        status=status.HTTP_400_BAD_REQUEST)

    # IMPORTANTE: Para las trasnferencias a cuentas polipay o a terceros (de lado de la wallet),el campo cuentatrasnferencia_id
    #               se establece el id de la cuenta del emisor, pero del lado de la banca no siempre se ocupa dicho
    #               campo, asi que se toma el campo cuenta_emisor. En caso de que si ocupen el campo y establezcan el id
    #               de la cuenta beneficiario para Recibidas, Solicitud de Saldos o Terceros a Polipay, se puede ocupar.

    objJson = {}
    objJson["tipo_pago"]                = queryTransferencia[0]["tipo_pago_id"]
    objJson["status_trans"]             = queryTransferencia[0]["status_trans_id"]
    objJson["monto"]                    = queryTransferencia[0]["monto"]
    if opcion == 3:
        objJson["nombre_emisor"]        = queryTransferencia[0]["empresa"]
    else:
        objJson["nombre_emisor"]        = queryTransferencia[0]["nombre_emisor"]
    objJson["nombre_beneficiario"]      = queryTransferencia[0]["nombre_beneficiario"]
    objJson["referencia_numerica"]      = queryTransferencia[0]["referencia_numerica"]
    if opcion == 3:
        objJson["cuentatransferencia"]  = pk_cuenta
    else:
        objJson["cuentatransferencia"]  = queryTransferencia[0]["cuentatransferencia_id"]
    objJson["id_transferencia"]         = queryTransferencia[0]["id"]

    # Escenario en el que se realicen los preparativos para RECIBIDAS
    if opcion == 2:
        notifyAppUserFromWeb(objJson, 3)
    elif opcion == 3:
        notifyAppUserFromWeb(objJson, 4)
    else:
        notifyAppUserFromWeb(objJson, 2)


def EstadoCuenta(username):
    ### pruebas
    masivoId = 767
    queryset = transferencia.objects.filter(masivo_trans_id=masivoId)
    periodo = transmasivaprod.objects.get(id=masivoId)
    query = queryset.last()
    cuentaUser = cuenta.objects.get(cuenta=query.cuenta_emisor)
    servicio = tipo_transferencia.objects.get(id=query.tipo_pago_id)
    usuario = persona.objects.get(id=cuentaUser.persona_cuenta_id)
    domicilioUser = domicilio.objects.get(domicilioPersona_id=usuario.id)
    ######################### 
    datos = 100
    numPagTotal = CalcularHojasTotales(datos)
    pdf_file = 'TMP/web/Estado_Cuentas/PDF/Estado-cuenta-' + username + ".pdf"
    carpetaImgEstadoCuenta = "TEMPLATES/web/EstadoCuenta/Img/"
    can = canvas.Canvas(pdf_file, pagesize=letter)
    can.drawImage(carpetaImgEstadoCuenta + "Estado-cuenta.jpg", 0, 660, width=612, height=125) #Logo
    can.drawImage(carpetaImgEstadoCuenta + "Informacion-periodo.jpg", 230, 585, width=230, height=60) #Informacion del periodo
    can.drawImage(carpetaImgEstadoCuenta + "Informacion-financiera.jpg", 33, 360, width=250, height=160)
    can.drawImage(carpetaImgEstadoCuenta + "Detalle-cargos.jpg", 301, 367, width=285, height=133.5)
    can.drawImage(carpetaImgEstadoCuenta + "Detalles-movimientos-realizados.jpg", 33, 260, width=547, height=80) #80
    can.drawImage(carpetaImgEstadoCuenta + "Pie-pagina.jpg", 33, 25, width=408, height=27)
    # can.setFontSize(8.5)
    # can.setFillColor("black")
    # can.drawString(44, 655, usuario.name)  # empresa
    # can.drawString(44, 642, servicio.nombre_tipo)  # nombre del servicio ejemplo Polipay Dispersa
    # can.setFillColor(HexColor('#5c5a59'))
    # can.drawString(44, 630, "No. Cuenta: " + str(cuentaUser.cuenta))  # NO. cuenta
    # can.drawString(44, 617, "No. Cliente / Prefijo: " + str(usuario.id))  # NO.CLIENTE/b PREFiJO
    # can.drawString(44, 605, "RFC: " + usuario.rfc)  # RFC
    # can.drawString(44, 585, domicilioUser.calle + " " + domicilioUser.no_exterior)  # Calle y numero
    # can.drawString(44, 570, domicilioUser.colonia)  # Colonia
    # can.drawString(44, 555, domicilioUser.alcaldia_mpio)  # Delegacion
    # can.drawString(44, 540, "CP " + domicilioUser.codigopostal + " " + domicilioUser.estado)  # CP y Estado
    # can.drawString(305, 607, periodo.observations)  # Periodo
    # can.drawString(305, 587, periodo.observations)  # Corte ------------------- pendiente ver como conseguirlo
    # can.drawString(410, 453, "Este producto no genera ninguna comisión")  # commissions Cobradas
    # can.drawString(410, 434, "No aplica")  # commissions Cobradas
    # can.drawString(410, 417, "En este período no se presentaron cargos objetados")
    can.save()
    return

# (Jose 2021/12/30) Funcion para detectar los meses
def Meses(mes):
    diccionario = {
        '01': 'Ene',
        '02': 'Feb',
        '03': 'Mar',
        '04': 'Abr',
        '05': 'May',
        '06': 'Jun',
        '07': 'Jul',
        '08': 'Ag',
        '09': 'Sept',
        '10': 'Oct',
        '11': 'Nov',
        '12': 'Dic'
    }
    mes = diccionario.get(mes)
    return mes