import shutil

from django.core.mail import message
from django.http.response import FileResponse
from django.db import transaction

from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet
from rest_framework.generics import ListAPIView
from rest_framework import status

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from MANAGEMENT.Standard.errors_responses import MyHttpError
from polipaynewConfig.inntec import listCard
from polipaynewConfig.exceptions import *
from apps.users.api.web.serializers.inntec_serializer import *
from apps.users.management import *
from apps.permision.permisions import BlocklistPermissionV2
import openpyxl
import mimetypes

class BuscarNumeroTarjetaCuentaEje(ListAPIView):

    def list(self, request, *args, **kwargs):
        person_id = get_Object_orList_error(persona, id=self.request.query_params["id"]).get_only_id()
        if grupoPersona.objects.filter(empresa_id=person_id):
            numero_tarjeta = self.request.query_params["Tarjeta"]
            tarjetas = tarjeta.objects.filter(clientePrincipal_id=person_id, status="04")
            cards = [i.get_tarjeta() for i in tarjetas if numero_tarjeta in i.get_tarjeta()]
            return Response({"status": {"tarjeta": cards}}, status=status.HTTP_200_OK)
        else:
            return Response({"status": "cuenta eje no encontrada"}, status=status.HTTP_400_BAD_REQUEST)


class BuscarNumeroTarjetaInntec(GenericViewSet):
    serializer_class = None

    def list(self, request):
        numero_tarjeta = self.request.query_params["Tarjeta"]
        data = listCard(numero_tarjeta)  # (Produccion)
        return Response(data, status=status.HTTP_200_OK)


class BuscarNumeroTarjetaInntecPrueba(GenericViewSet):
    serializer_class = None

    def list(self, request):
        numero_tarjeta = self.request.query_params["Tarjeta"]
        data = listCardPrueba(numero_tarjeta) #(Prueba)
        return Response(data, status=status.HTTP_200_OK)


class AsignarTarjetaInntecPersonalExterno(GenericViewSet):
    # permission_classes = (BlocklistPermissionV2,)
    # permisos = ["Asignar tarjeta a personal externo por grupo", "Asignar tarjeta a personal externo por centro de costo"]
    serializer_class = SerializerAsignarTarjetasPersonaExterna
    permission_classes = ()

    def create(self):
        pass

    def put(self, request):
        instance = get_Object_orList_error(cuenta, cuenta=request.data["cuenta"])
        queryset = grupoPersona.objects.filter(person_id=instance.persona_cuenta_id, relacion_grupo_id__in=[6, 9])
        if len(queryset) != 0:
            pk_empresa = request.data["cuentaEje"]
            pk_cuenta_eje = grupoPersona.objects.filter(empresa_id=pk_empresa)
            if len(pk_cuenta_eje) != 0:
                serializer = self.serializer_class(data=request.data, context={"empresa_id": pk_empresa})
                if serializer.is_valid(raise_exception=True):
                    try:
                        with transaction.atomic():
                            instanceP = get_Object_orList_error(persona, id=instance.persona_cuenta_id)
                            tarjetas = serializer.update(instance, instanceP)
                            return Response({"status": {"Se_asignaron_las_tarjetas": tarjetas}}, status=status.HTTP_200_OK)
                    except Exception as e:
                        message = "Ocurrio un error durante el proceso de de asignar tarjeta, Error:   " + str(e)
                        error = {'field':'', "data":'', 'message': message}
                        MensajeError(error)

            else:
                return Response({"status": "Cuenta eje no encontrada"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"status": "persona externa no encontrada"}, status=status.HTTP_400_BAD_REQUEST)


class AsignarTarjetaInnteCuentaEje(GenericViewSet):
    # permission_classes = (BlocklistPermissionV2,)
    # permisos = ["Asignar stock de tarjetas por centro de costo"]
    serializer_class = SerializerAsignarTarjetasCuentaEje
    permission_classes = ()


    def create(self, request):
        instance = persona.objects.get(id=self.request.query_params["id"])
        if grupoPersona.objects.filter(empresa_id=instance.id):
            serializer = self.serializer_class(data=request.data)
            tarjetas = request.data["tarjeta"]
            tarjetas,datos_tarjeta_inntec,token = serializer.validar_tarjetas(tarjetas)
            try:
                with transaction.atomic():
                    tarjetas = serializer.create(instance,tarjetas,datos_tarjeta_inntec,token)
                    return Response({"status": {"Se_asignaron_las_tarjetas": tarjetas}}, status=status.HTTP_200_OK)
            except Exception as e:
                message = "Ocurrio un error durante el proceso de de asignar tarjetas, Error:   " + str(e)
                error = {'field':'', "data":'', 'message': message}
                MensajeError(error)
        else:
            return Response({"status": "cuenta eje no encontrada"}, status=status.HTTP_400_BAD_REQUEST)


class AsignarTarjetaInnteCuentaEjePrueba(GenericViewSet): ##########################
    # permission_classes = (BlocklistPermissionV2,)
    # permisos = ["Asignar stock de tarjetas por centro de costo"]
    serializer_class = SerializerAsignarTarjetasCuentaEjePrueba
    permission_classes = ()

    def create(self, request):
        instance = get_Object_orList_error(persona, id=self.request.query_params["id"])
        if grupoPersona.objects.filter(empresa_id=instance.id):
            serializer = self.serializer_class(data=request.data)
            tarjetas = request.data["tarjeta"]
            tarjetas,datos_tarjeta_inntec,token = serializer.validar_tarjetas(tarjetas)
            try:
                with transaction.atomic():
                    tarjetas = serializer.create(instance,tarjetas,datos_tarjeta_inntec,token)
                    return Response({"status": {"Se_asignaron_las_tarjetas": tarjetas}}, status=status.HTTP_200_OK)
            except Exception as e:
                message = "Ocurrio un error durante el proceso de de asignar tarjetas, Error:   " + str(e)
                error = {'field':'', "data":'', 'message': message}
                MensajeError(error)
        else:
            return Response({"status": "cuenta eje no encontrada"}, status=status.HTTP_400_BAD_REQUEST)


class ComponentExportExcelBeneficiario:
    xlsx_beneficiario: ClassVar[str] = "TEMPLATES/web/Excel-beneficiarios-sin-tarjeta/Layout-Beneficiarios.xlsx"
    absolute_url: ClassVar[str] = "TMP/web/Layour-ben-sin-tarjetas"
    cell_position_default: ClassVar[int] = 2

    def __init__(self, list_beneficiarios: List[Dict[str, Any]]):
        dt = datetime.datetime.strftime(datetime.datetime.now(), 'Y%m%d%H%M')
        self.filename = f"{self.absolute_url}/LayoutBeneficiarios_{dt}.xlsx"
        # (2022.06.01 - ChrAvaBus) Se modifica para que cree el archivo antes de descargarse
        print("Antes de crear...")
        archivoOrigen   = str(self.xlsx_beneficiario)
        archivoDestino  = str(self.filename)
        print("origen["+str(archivoOrigen)+"]")
        print("origen[" + str(archivoDestino) + "]")
        shutil.copyfile(archivoOrigen, archivoDestino)
        print("Despues de crear...")

        self.list_beneficiarios = list_beneficiarios
        self.workbook = load_workbook(filename=self.xlsx_beneficiario)
        self.sheet = self.workbook.active
        self.write_xlsx()
        self.save()

        self.response = self.response_file(self.filename)

    def save(self):
        self.workbook.save(filename=self.filename)

    @staticmethod
    def response_file(filename: str):
        response = FileResponse(open(filename, 'rb'))
        response['Content-Disposition'] = "attachment; filename=%s" % filename
        os.remove(filename)
        return response

    def write_xlsx(self):
        cell_position_default = self.cell_position_default
        for row in self.list_beneficiarios:
            self.sheet['A' + str(cell_position_default)] = row.get('persona_cuenta__name').title()

            try:
                apellido_paterno, apellido_materno = (row.get('persona_cuenta__last_name')).split("*")
            except Exception as e:
                apellido_paterno, apellido_materno = row.get('persona_cuenta__last_name'), ''

            self.sheet['B' + str(cell_position_default)] = apellido_paterno.title()
            self.sheet['C' + str(cell_position_default)] = apellido_materno.title()
            self.sheet['D' + str(cell_position_default)] = row.get('cuenta')
            cell_position_default += 1


# Esta con API de pruebas
class LayoutAsignarTarjetaPersonalExterno(GenericViewSet):
    permission_classes = ()
    serializer_class = AsignarTarjetasBeneficiarioMasivo

    @staticmethod
    def list_beneficiario(razon_social_id: int) -> List[int]:
        return grupoPersona.objects.filter(
            empresa_id=razon_social_id,
            relacion_grupo_id=6).values_list(
            'person_id',
            flat=True
        )

    @staticmethod
    def list_info_beneficiario(person_list: List[int]) -> List[Dict[str, Any]]:
        return cuenta.objects.filter(persona_cuenta_id__in=person_list).values(
            'id',
            'persona_cuenta__name',
            'persona_cuenta__last_name',
            'cuenta'
        )

    def list(self, request):
        try:
            beneficiarios_list = self.list_beneficiario(self.request.query_params["CuentaEjeId"])
            # (2022.06.01 - ChrAvaBus) Se modifica para que regrese el excel vacio, aun sin tener beneficiarios
            if not beneficiarios_list:
                #raise ValueError('No hay información por mostrar')
                export_xls = ComponentExportExcelBeneficiario([])
                return export_xls.response

            beneficiarios_list_info = self.list_info_beneficiario(beneficiarios_list)
            export_xls = ComponentExportExcelBeneficiario(beneficiarios_list_info)
            return export_xls.response
        except ValueError as e:
            err = MyHttpError(message=str(e), real_error=str(e))
            return Response(err.standard_error_responses(), status=status.HTTP_400_BAD_REQUEST)

        # for persona_externa in g:
        #     try:
        #         cuenta_persona_externa = cuenta.objects.get(persona_cuenta_id=persona_externa.get('person_id'))
        #         if tarjeta.objects.filter(cuenta_id=cuenta_persona_externa.id):
        #             persona_externa_sin_tarjeta = persona.objects.get(id=persona_externa.get('person_id'))
        #             usuarios_sin_tarjeta.append(persona_externa_sin_tarjeta)
        #         else:
        #             persona_externa_sin_tarjeta = persona.objects.get(id=persona_externa.get('person_id'))
        #             usuarios_sin_tarjeta.append(persona_externa_sin_tarjeta)
        #     except TypeError as e:
        #         print(e)
        #         continue

        # Conseguir las personas sin tarjeta
        # Abrir excel
        # excel_beneficiarios_sin_tarjetas = load_workbook(filename=self.xlsx_beneficiario)
        # sheet = excel_beneficiarios_sin_tarjetas.active
        # # Colocar datos
        # numero = 2
        # for datos in beneficiarios_list_info:
        #     sheet['A' + str(numero)] = datos.get('persona_cuenta__name').title()
        #     try:
        #         apellido_paterno, apellido_materno = (datos.get('persona_cuenta__last_name')).split("*")
        #     except:
        #         apellido_paterno, apellido_materno = datos.get('persona_cuenta__last_name'), ''
        #     sheet['B' + str(numero)] = apellido_paterno.title()
        #     sheet['C' + str(numero)] = apellido_materno.title()
        #     # numero_cuenta = cuenta.objects.get(persona_cuenta_id=datos.id)
        #     sheet['D' + str(numero)] = datos.get('cuenta')
        #     numero = numero + 1
        #
        # fecha_actual = str(datetime.date.today()).replace('-', '')
        # excel_beneficiarios_sin_tarjetas.save(filename=f"{self.filename}_{username_excel}.xlsx")
        # # Descargar Excel
        # filename = f"{self.filename}_{username_excel}.xlsx"
        # filepath = filename
        # # path = open(filepath, 'r')
        # # mime_type, _ = mimetypes.guess_type(filepath)
        # response = FileResponse(open(filename, 'rb'))
        # response['Content-Disposition'] = "attachment; filename=%s" % filename
        # os.remove(filename)
        # return response

    def create(self, request):
        cuenta_eje = self.request.query_params['CuentaEjeId']
        listado_excel = request.data["PersonList"]
        estado = request.query_params["Estado"]
        self.serializer_class.validate_tarjeta(listado_excel,cuenta_eje,estado)
        try:
            with transaction.atomic():
                self.serializer_class.Asignar_tarjetas(listado_excel,estado)
                mensaje,data,field = "Se asignaron las tarjetas a los beneficiarios", "Null", "Null"
                respuesta = MessageOK(mensaje,data,field)
                return Response(respuesta, status=status.HTTP_200_OK)
        except Exception as e:
            message = "Ocurrio un error durante el proceso de creación de Beneficiarios masivos, Error:   " + str(e)
            error = {'field':'', "data":'', 'message': message}
            MensajeError(error)





class CheckCardBalance(ListAPIView):
    permission_classes = ()

    def list(self, request, *args, **kwargs):
        card_number = request.query_params['card_number']
        inntec_balance = check_inntec_card_balance(card_number)
        return Response(inntec_balance, status=status.HTTP_200_OK)

# Prueba de dar de alta la tarjeta del beneficiario
class PruebaDeTarjeta(GenericViewSet):

    def create(self,request):
        Tarjeta = request.data["tarjeta"]
        cuenta_eje_id = request.data["cuentaId"]
        beneficiario = request.data["beneficiario"]
        TarjetaBeneficiario(Tarjeta,cuenta_eje_id,beneficiario)
        return Response({"status": "Se asigno la tarjeta"}, status=status.HTTP_200_OK)




